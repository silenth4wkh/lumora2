#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Részletes Excel fájl ellenőrzés"""
import openpyxl
from glob import glob
import os

excel_files = glob('it_allasok_*.xlsx')
if not excel_files:
    print("[ERROR] Nincs Excel fajl")
    exit(1)

latest_file = max(excel_files, key=os.path.getctime)
print(f"Excel fajl: {os.path.basename(latest_file)}")
print("="*60)

wb = openpyxl.load_workbook(latest_file)

# Portálok számlálása
portals = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    print(f"  Sorok: {ws.max_row - 1} (fejlec nelkul)")
    
    # Portálok számlálása
    if ws.max_row > 1:
        source_col = None
        for col_idx, header in enumerate(ws[1], 1):
            if header.value and ('forrás' in str(header.value).lower() or 'forras' in str(header.value).lower()):
                source_col = col_idx
                break
        
        if source_col:
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=source_col).value
                if cell_value:
                    source = str(cell_value).split(' – ')[0] if ' – ' in str(cell_value) else str(cell_value).split(' - ')[0]
                    portals[source] = portals.get(source, 0) + 1
            
            print(f"  Portálok:")
            for portal, count in sorted(portals.items(), key=lambda x: x[1], reverse=True):
                print(f"    - {portal}: {count} allas")

# Összesítés
print("\n" + "="*60)
print("OSSZESITES:")
print(f"  Osszes allas: {sum(portals.values())}")
print(f"  Portálok szama: {len(portals)}")
print("="*60)

