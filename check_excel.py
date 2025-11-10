#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excel fájl tartalmának ellenőrzése"""
import openpyxl
import os
from glob import glob

# Legutóbbi Excel fájl keresése
excel_files = glob('it_allasok_*.xlsx')
if not excel_files:
    print("[ERROR] Nincs Excel fajl")
    exit(1)

latest_file = max(excel_files, key=os.path.getctime)
print(f"Excel fajl: {latest_file}")
print(f"Meret: {os.path.getsize(latest_file) / 1024:.2f} KB")
print("="*60)

try:
    wb = openpyxl.load_workbook(latest_file)
    print(f"\nSheet-ek szama: {len(wb.sheetnames)}")
    print(f"Sheet nevek: {', '.join(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} Sheet ---")
        print(f"Max sor: {ws.max_row}")
        print(f"Max oszlop: {ws.max_column}")
        
        if ws.max_row > 1:
            # Fejléc
            headers = [cell.value for cell in ws[1]]
            print(f"Fejlec: {headers[:5]}...")
            
            # Első néhány sor
            print("\nElso 5 sor adatai:")
            for row_num in range(2, min(7, ws.max_row + 1)):
                row_data = [str(cell.value)[:30] if cell.value else "" for cell in ws[row_num]]
                print(f"  Sor {row_num-1}: {row_data[0]} | {row_data[1][:20]} | {row_data[2][:20]} | {row_data[3][:20]}")
        else:
            print("  [FIGYELEM] Nincs adat a sheet-ben!")
    
    print("\n" + "="*60)
    print("Excel fajl ellenorzes kesz!")
    
except Exception as e:
    print(f"[ERROR] Hiba az Excel fajl olvasasakor: {e}")
    import traceback
    traceback.print_exc()

