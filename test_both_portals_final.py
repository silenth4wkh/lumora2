#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teljes teszt: Profession.hu + No Fluff Jobs együtt
"""
import requests
import time
import openpyxl
from datetime import datetime

print("=" * 80)
print("PROFESSION.HU + NO FLUFF JOBS EGYÜTTES TESZT")
print("=" * 80)

# 1. Scraping mindkét portálról
print("\n[1] Scraping indítása (mindkét portál)...")
time.sleep(2)

start_time = datetime.now()
r = requests.post('http://127.0.0.1:5000/api/search', json={
    'categories': ['IT']  # Ez az egyetlen szükséges paraméter
}, timeout=1800)  # 30 perc timeout

end_time = datetime.now()
duration = (end_time - start_time).total_seconds()

print(f"Status: {r.status_code}")
print(f"Időtartam: {duration:.1f} másodperc")

data = r.json()
print(f"\nTalált állások összesen: {data.get('total_jobs')}")
print(f"Üzenet: {data.get('message')}")

# Portálonkénti statisztika
jobs = data.get('jobs', [])
profession_count = sum(1 for j in jobs if 'Profession' in j.get('Forrás', ''))
nofluff_count = sum(1 for j in jobs if 'No Fluff' in j.get('Forrás', ''))

print(f"\nPortálonkénti bontás:")
print(f"  - Profession.hu: {profession_count}")
print(f"  - No Fluff Jobs: {nofluff_count}")

# 2. Példa állások mindkét portálról
print(f"\n[2] Példa állások:")
print("\nProfession.hu (első 3):")
for i, job in enumerate([j for j in jobs if 'Profession' in j.get('Forrás', '')][:3], 1):
    print(f"  {i}. {job.get('Pozíció', 'N/A')[:50]} | {job.get('Cég', 'N/A')[:30]}")

print("\nNo Fluff Jobs (első 3):")
for i, job in enumerate([j for j in jobs if 'No Fluff' in j.get('Forrás', '')][:3], 1):
    print(f"  {i}. {job.get('Pozíció', 'N/A')[:50]} | {job.get('Cég', 'N/A')[:30]} | Bérsáv: {job.get('Fizetés', 'N/A')[:30]}")

# 3. Excel export
print(f"\n[3] Excel export...")
r2 = requests.post('http://127.0.0.1:5000/api/export/excel', json={}, timeout=60)

if r2.status_code == 200:
    filename = f"it_allasok_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    with open(filename, 'wb') as f:
        f.write(r2.content)
    print(f"[OK] Excel letöltve: {filename}")
    
    # 4. Excel ellenőrzés
    print(f"\n[4] Excel tartalom...")
    wb = openpyxl.load_workbook(filename)
    
    print(f"Sheet-ek: {wb.sheetnames}")
    
    if len(wb.sheetnames) > 1:
        # Multi-portal: külön sheet-ek
        print("\n[Multi-portal export]")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name != "Összesítés":
                rows = ws.max_row - 1
                print(f"  - {sheet_name}: {rows} állás")
    else:
        # Single-portal: egy sheet
        ws = wb.active
        rows = ws.max_row - 1
        print(f"\n[Single-portal export]")
        print(f"  - Sorok: {rows}")
    
    # Példa sorok
    ws = wb.active if len(wb.sheetnames) == 1 else wb[wb.sheetnames[1]]
    print(f"\nPélda sorok (első sheet):")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
        print(f"  {i}. {row[1][:20]} | {row[2][:40]} | Publikálva: {row[8]}")
    
    print("\n" + "=" * 80)
    print("[SIKERES TESZT]")
    print(f"Összesen: {profession_count + nofluff_count} állás")
    print(f"  - Profession.hu: {profession_count}")
    print(f"  - No Fluff Jobs: {nofluff_count}")
    print(f"Időtartam: {duration:.1f}s")
    print("=" * 80)

else:
    print(f"[HIBA] Excel export failed: {r2.status_code}")
    print(r2.text)

