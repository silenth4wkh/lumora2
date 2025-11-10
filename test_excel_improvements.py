#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Test Excel improvements: Leírás, Bérsáv, Lekérés dátuma
"""
import requests
import time
import openpyxl
from datetime import datetime, timedelta

print("=" * 80)
print("EXCEL FEJLESZTÉSEK TESZTJE")
print("=" * 80)

# 1. Scraping
print("\n[1] Scraping...")
time.sleep(4)
r = requests.post('http://127.0.0.1:5000/api/search/nofluff-only', json={}, timeout=60)
print(f"Status: {r.status_code}")
data = r.json()
print(f"Jobs: {data.get('total_jobs')}")

# 2. Várj egy kicsit (hogy látható legyen az időkülönbség)
print("\n[2] Várunk 3 másodpercet...")
time.sleep(3)

# 3. Excel export
print("\n[3] Excel export...")
export_time_before = datetime.now()
r2 = requests.post('http://127.0.0.1:5000/api/export/excel', json={}, timeout=30)
export_time_after = datetime.now()

filename = f"it_allasok_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
with open(filename, 'wb') as f:
    f.write(r2.content)
print(f"Excel letöltve: {filename}")
print(f"Export idő: {export_time_before.strftime('%Y-%m-%d %H:%M:%S')}")

# 4. Excel ellenőrzés
print("\n[4] Excel tartalom ellenőrzése...")
wb = openpyxl.load_workbook(filename)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"Fejlécek: {headers}")

print("\n[5] Első 5 állás részletesen:")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
    print(f"\n  Sor #{i}:")
    print(f"    Pozíció: {row[2][:50] if row[2] else 'N/A'}")
    print(f"    Cég: {row[3]}")
    print(f"    Bérsáv (Fizetés): {row[5]}")  # col 6 = Fizetés
    print(f"    Publikálva: {row[8]}")  # col 9 = Publikálva
    print(f"    Lekérés dátuma: {row[9]}")  # col 10 = Lekérés dátuma
    print(f"    Leírás: {row[10][:50] if row[10] else '(üres)'}...")  # col 11 = Leírás

# 6. Validálás
print("\n[6] Validálás:")
success = True

# Bérsáv ellenőrzés
salary_count = sum(1 for row in ws.iter_rows(min_row=2, max_row=20, values_only=True) if row[5])
print(f"  - Bérsávval rendelkezők (első 20-ból): {salary_count}/19")

# Lekérés dátuma == ma
export_date_expected = datetime.today().strftime('%Y-%m-%d')
export_dates = [row[9] for row in ws.iter_rows(min_row=2, max_row=20, values_only=True)]
correct_dates = sum(1 for d in export_dates if d == export_date_expected)
print(f"  - Lekérés dátuma == {export_date_expected}: {correct_dates}/19")

if salary_count > 10 and correct_dates == 19:
    print("\n" + "=" * 80)
    print("[SIKERES] Minden fejlesztés működik!")
    print("=" * 80)
else:
    print("\n" + "=" * 80)
    print("[FIGYELEM] Ellenőrizd az eredményeket")
    print("=" * 80)

