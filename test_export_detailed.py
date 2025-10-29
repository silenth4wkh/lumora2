"""Részletes teszt az Excel export problémájának felderítésére"""
import requests
import openpyxl
from io import BytesIO

print("1. Scraper futtatása...")
response = requests.post('http://localhost:5000/api/test/nofluffjobs-no-dedup', json={})
data = response.json()

print(f"   Scraper eredmény: {data.get('count', 0)} állás")
print(f"   Cache updated: {data.get('cache_updated', False)}")

if 'jobs' in data:
    print(f"\n2. Első állás struktúrája:")
    first_job = data['jobs'][0]
    print(f"   Mezők: {list(first_job.keys())}")
    for key, value in first_job.items():
        print(f"   {key}: {value[:80] if isinstance(value, str) and value else value}")

print("\n3. Excel export...")
response = requests.get('http://localhost:5000/api/export/excel')

print(f"   Response status: {response.status_code}")
print(f"   Content length: {len(response.content)} bytes")

# Excel fájl megnyitása memóriában
try:
    wb = openpyxl.load_workbook(BytesIO(response.content))
    print(f"   Excel megnyitva: OK")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n4. Sheet: {sheet_name}")
        print(f"   Sorok: {ws.max_row}")
        print(f"   Oszlopok: {ws.max_column}")
        
        # Első sor (fejléc)
        print(f"\n   Fejléc sor:")
        for col in range(1, min(13, ws.max_column + 1)):
            val = ws.cell(row=1, column=col).value
            print(f"      [{col}] {val}")
        
        # Néhány adat sor ellenőrzése
        print(f"\n   Adat sorok (első 3):")
        for row in range(2, min(5, ws.max_row + 1)):
            print(f"      Sor {row}:")
            for col in range(1, min(13, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val:
                    print(f"         [{col}] {str(val)[:50]}")
        
        # Üres sorok száma
        empty_rows = 0
        for row in range(2, ws.max_row + 1):
            has_data = False
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    break
            if not has_data:
                empty_rows += 1
        
        print(f"\n   Üres sorok: {empty_rows}/{ws.max_row - 1}")
        
except Exception as e:
    print(f"   Hiba az Excel megnyitásakor: {e}")
    import traceback
    traceback.print_exc()

print("\nKÉSZ!")

