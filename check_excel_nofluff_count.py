#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Check how many No Fluff Jobs are in the latest Excel file"""
import os, glob
from openpyxl import load_workbook

def main():
    # Find latest Excel file
    excel_files = glob.glob("it_allasok_render_*.xlsx") + glob.glob("it_allasok_*.xlsx")
    if not excel_files:
        print("[ERROR] No Excel files found")
        return
    
    latest = max(excel_files, key=os.path.getmtime)
    print(f"[CHECKING] {latest}")
    
    wb = load_workbook(latest)
    
    print(f"\n[EXCEL] Sheets: {wb.sheetnames}")
    
    total_nofluff = 0
    total_profession = 0
    total_other = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[{sheet_name}]")
        print(f"  Rows: {ws.max_row - 1} (excluding header)")
        
        nofluff = 0
        profession = 0
        other = 0
        
        # Check Forrás column (B column, index 2)
        for row in range(2, ws.max_row + 1):
            source_cell = ws.cell(row, 2)  # B column
            source = str(source_cell.value or "").lower()
            
            if "no fluff" in source or "nofluff" in source:
                nofluff += 1
            elif "profession" in source:
                profession += 1
            else:
                other += 1
        
        print(f"  - No Fluff Jobs: {nofluff}")
        print(f"  - Profession.hu: {profession}")
        print(f"  - Other: {other}")
        
        total_nofluff += nofluff
        total_profession += profession
        total_other += other
    
    print("\n" + "="*70)
    print("[SUMMARY]")
    print(f"Total No Fluff Jobs: {total_nofluff}")
    print(f"Total Profession.hu: {total_profession}")
    print(f"Total Other: {total_other}")
    print(f"Grand Total: {total_nofluff + total_profession + total_other}")
    print("="*70)
    
    if total_nofluff < 100:
        print(f"\n[WARNING] Only {total_nofluff} No Fluff Jobs - expected ~795")
        print("          Possible issues:")
        print("          1. API not used (HTML fallback only)")
        print("          2. Duplication filtering too aggressive")
        print("          3. Excel export filtering issue")
    else:
        print(f"\n[OK] {total_nofluff} No Fluff Jobs found")

if __name__ == '__main__':
    main()

