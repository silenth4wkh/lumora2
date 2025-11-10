#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel export ellenőrzése és letöltése
"""

import requests
import openpyxl
from io import BytesIO
from datetime import datetime
import os

BASE_URL = "http://127.0.0.1:5000"

def check_excel_export():
    """Excel export ellenőrzése"""
    print("=" * 60)
    print("EXCEL EXPORT ELLENŐRZÉSE")
    print("=" * 60)
    print()
    
    try:
        # Excel export lekérése
        print("[1] Excel export lekérése...")
        response = requests.get(f"{BASE_URL}/api/export/excel", timeout=30)
        
        if response.status_code == 200:
            print("[OK] Excel export sikeres")
            print(f"    Fájlméret: {len(response.content) / 1024:.2f} KB")
            print()
            
            # Excel fájl mentése
            filename = f"it_allasok_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            with open(filename, 'wb') as f:
                f.write(response.content)
            
            print(f"[OK] Excel fájl mentve: {filename}")
            print(f"    Teljes útvonal: {os.path.abspath(filename)}")
            print()
            
            # Excel tartalom ellenőrzése
            print("[2] Excel tartalom ellenőrzése...")
            
            wb = openpyxl.load_workbook(BytesIO(response.content))
            
            print(f"    Sheet-ek száma: {len(wb.sheetnames)}")
            print(f"    Sheet nevek: {', '.join(wb.sheetnames)}")
            print()
            
            # Minden sheet ellenőrzése
            total_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.max_row - 1  # -1 a fejléc miatt
                cols = ws.max_column
                
                print(f"    [{sheet_name}]")
                print(f"      Sorok (adat): {rows}")
                print(f"      Oszlopok: {cols}")
                
                total_rows += rows
                
                # Portál szerinti bontás
                if rows > 0:
                    nofluff_count = 0
                    profession_count = 0
                    
                    for row in range(2, min(ws.max_row + 1, 1002)):  # Max 1000 sor ellenőrzés
                        source_cell = ws.cell(row, 2)  # B oszlop = Forrás
                        if source_cell.value:
                            source = str(source_cell.value).lower()
                            if "no fluff" in source:
                                nofluff_count += 1
                            elif "profession" in source:
                                profession_count += 1
                    
                    print(f"      No Fluff Jobs (első 1000 sorban): ~{nofluff_count}")
                    print(f"      Profession.hu (első 1000 sorban): ~{profession_count}")
                
                print()
            
            print(f"[STATS] Összesen {total_rows} állás az Excel-ben")
            
            return filename
        else:
            print(f"[HIBA] Excel export hiba: {response.status_code}")
            print(f"    Válasz: {response.text[:200]}")
            return None
        
    except Exception as e:
        print(f"[HIBA] {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    check_excel_export()

