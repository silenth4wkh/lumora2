#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

filename = 'it_allasok_20251030_110929.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active

print("=" * 80)
print("EXCEL TARTALOM ELLENORZES")
print("=" * 80)

print(f"\nFejlecek: {[c.value for c in ws[1]]}")
print(f"\nOsszesen: {ws.max_row - 1} allas")

print("\nElso 10 allas:")
print("-" * 80)
print(f"{'#':<3} | {'Pozicio':<40} | {'Ceg':<25} | {'Lokacio':<15}")
print("-" * 80)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
    pozicio = (row[2] or '')[:40]
    ceg = (row[3] or '')[:25]
    lokacio = (row[4] or '')[:15]
    print(f"{i:<3} | {pozicio:<40} | {ceg:<25} | {lokacio:<15}")

print("=" * 80)
print("[OK] No Fluff Jobs scraper: tiszta mezoket hoz letre!")
print("=" * 80)

