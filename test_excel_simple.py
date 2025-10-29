import requests
import openpyxl
from io import BytesIO

# Scraper futtatása
print("1. No Fluff Jobs scraper...")
r1 = requests.post('http://localhost:5000/api/test/nofluffjobs-no-dedup', json={})
print(f"   Status: {r1.status_code}, Count: {r1.json().get('count', 0)}")

# Excel export
print("\n2. Excel export...")
r2 = requests.get('http://localhost:5000/api/export/excel')
print(f"   Status: {r2.status_code}, Size: {len(r2.content)} bytes")

# Excel beolvasása
wb = openpyxl.load_workbook(BytesIO(r2.content))
ws = wb.active

print(f"\n3. Excel tartalom:")
print(f"   Sorok: {ws.max_row}, Oszlopok: {ws.max_column}")
print(f"   Első 10 sor tartalma:")

for row in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(6, ws.max_column + 1)):
        val = ws.cell(row=row, column=col).value
        row_data.append(str(val)[:30] if val else "")
    print(f"   Sor {row}: {' | '.join(row_data)}")

