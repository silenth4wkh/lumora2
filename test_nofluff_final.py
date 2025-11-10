#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teljes No Fluff Jobs teszt: scraping → Excel export ellenőrzés
"""
import requests
import time
import openpyxl

print("=" * 60)
print("NO FLUFF JOBS TELJES TESZT")
print("=" * 60)

# 1. Scraping
print("\n[1] Scraping indítása...")
url = 'http://127.0.0.1:5000/api/search/nofluff-only'
try:
    time.sleep(3)  # Flask app startup
    r = requests.post(url, json={}, timeout=600)
    r.raise_for_status()
    data = r.json()
    print(f"[OK] Status: {r.status_code}")
    print(f"[OK] Talált állások: {data.get('total_jobs')}")
    jobs = data.get('jobs', [])
    
    if jobs:
        print(f"\n[2] Első 3 állás ellenőrzése:")
        for i, job in enumerate(jobs[:3], 1):
            print(f"\n  Állás #{i}:")
            print(f"    Pozíció: {job.get('Pozíció', 'N/A')[:60]}")
            print(f"    Cég: {job.get('Cég', 'N/A')}")
            print(f"    Lokáció: {job.get('Lokáció', 'N/A')}")
            print(f"    Publikálva: {job.get('Publikálva', 'N/A')}")
            print(f"    Link: {job.get('Link', 'N/A')[:50]}...")
    
    # 2. Excel export
    print(f"\n[3] Excel export kérése...")
    r2 = requests.post('http://127.0.0.1:5000/api/export/excel', json={}, timeout=30)
    r2.raise_for_status()
    
    # Excel fájl mentése
    import re
    from datetime import datetime
    filename = f"it_allasok_{datetime.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with open(filename, 'wb') as f:
        f.write(r2.content)
    print(f"[OK] Excel fájl letöltve: {filename}")
    
    # 3. Excel fájl ellenőrzése
    print(f"\n[4] Excel fájl tartalmának ellenőrzése...")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"[OK] Fejlécek: {headers}")
    
    row_count = ws.max_row - 1
    print(f"[OK] Sorok száma (fejléc nélkül): {row_count}")
    
    print(f"\n[5] Első 5 sor részletes ellenőrzése:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
        print(f"\n  Sor #{i}:")
        print(f"    ID: {row[0]}")
        print(f"    Forrás: {row[1]}")
        print(f"    Pozíció: {row[2][:60] if row[2] else 'N/A'}")
        print(f"    Cég: {row[3]}")
        print(f"    Lokáció: {row[4]}")
        print(f"    Publikálva: {row[8]}")
    
    print("\n" + "=" * 60)
    print("[TESZT SIKERES]")
    print("=" * 60)
    
except Exception as e:
    print(f"\n[HIBA] {e}")
    import traceback
    traceback.print_exc()

