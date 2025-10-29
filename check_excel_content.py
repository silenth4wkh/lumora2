"""Excel fájl tartalmának ellenőrzése"""
import openpyxl
import sys

def check_excel_content(filename):
    """Excel fájl tartalmának ellenőrzése"""
    try:
        # Excel fájl megnyitása
        wb = openpyxl.load_workbook(filename)
        
        print(f"Excel fájl megnyitva: {filename}")
        print(f"Sheet-ek száma: {len(wb.sheetnames)}")
        print(f"Sheet nevek: {wb.sheetnames}\n")
        
        # Minden sheet ellenőrzése
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"=== Sheet: {sheet_name} ===")
            print(f"Sorok száma: {ws.max_row}")
            print(f"Oszlopok száma: {ws.max_column}\n")
            
            # Fejléc sor
            print("Fejléc sor:")
            for col in range(1, min(ws.max_column + 1, 13)):
                cell = ws.cell(row=1, column=col)
                print(f"  [{col}] {cell.value}")
            print()
            
            # Első néhány adat sor
            print("Adat sorok (első 5):")
            for row in range(2, min(ws.max_row + 1, 7)):
                print(f"  Sor {row}:")
                for col in range(1, min(ws.max_column + 1, 6)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        print(f"    [{col}] {str(cell.value)[:50]}")
                print()
            
            print(f"Összes sor: {ws.max_row}")
            print()
            
    except Exception as e:
        print(f"Hiba: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Teszt fájlok ellenőrzése
    test_files = [
        "nofluffjobs_debug.xlsx",
        "nofluffjobs_results.xlsx"
    ]
    
    for filename in test_files:
        print(f"\n{'='*60}")
        print(f"Ellenőrzés: {filename}")
        print(f"{'='*60}\n")
        check_excel_content(filename)

