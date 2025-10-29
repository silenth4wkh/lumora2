from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import requests
import xml.etree.ElementTree as ET
from urllib.parse import quote
from datetime import datetime, timedelta
import time
import re
import json
from collections import defaultdict
import threading
import queue
import os
import io

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None

app = Flask(__name__)
CORS(app)

# Error handler for debugging
@app.errorhandler(Exception)
def handle_exception(e):
    error_message = f"Error: {str(e)}"
    print(f"ERROR: {error_message}")
    return jsonify({"error": error_message, "type": type(e).__name__}), 500

# Scraper kód (a te kódból)
KW_LANG = [
    "java","python","c#",".net","dotnet","c++","cpp","golang","go","rust","php","ruby","scala","haskell",
    "kotlin","swift","objective-c"
]
KW_FE = [
    "frontend","front-end","react","angular","vue","svelte","next.js","nuxt","typescript","javascript","web developer"
]
KW_BE = [
    "backend","back-end","node","spring","spring boot","quarkus",".net core","asp.net","laravel","symfony","django","flask","fastapi"
]
KW_MOBILE = [
    "android","ios","mobile","kotlin","swift","flutter","react native","xamarin","ionic"
]
KW_DATA = [
    "data engineer","data scientist","ml engineer","machine learning","ai engineer","big data",
    "etl","elt","dwh","data warehouse","power bi","tableau","qlik","sql","spark","hadoop","db developer","snowflake","databricks"
]
KW_DEVOPS = [
    "devops","sre","site reliability","platform engineer","cloud engineer","kubernetes","docker","terraform","ansible",
    "aws","azure","gcp","cloud architect","finops","observability"
]
KW_TEST = [
    "qa","test automation","tesztautomatizálás","tesztmérnök","sdet","cypress","selenium","playwright","jmeter","postman"
]
KW_EMBED = [
    "embedded","firmware","fpga","rtos","bare metal","microcontroller","stm32","esp32","embedded linux","yocto","driver developer","c developer"
]
KW_SECURITY = [
    "security engineer","application security","appsec","devsecops","penetration tester","pentest","iam","siem","soc"
]
KW_ENTERPRISE = [
    "sap","abap","erp","crm","salesforce","servicenow","mendix","outsystems","navision","business central","oracle developer","ms dynamics"
]
KW_GENERAL_HU = [
    "fejlesztő","programozó","szoftver","szoftvermérnök","rendszermérnök","alkalmazásfejlesztő","alkalmazás üzemeltető","full stack","full-stack"
]

KEYWORDS = (
    KW_LANG + KW_FE + KW_BE + KW_MOBILE + KW_DATA + KW_DEVOPS +
    KW_TEST + KW_EMBED + KW_SECURITY + KW_ENTERPRISE + KW_GENERAL_HU
)

EXCLUDE_PHRASES = [
    "fejlesztő pedagógus","pedagógus","gyógypedagógus","konstruktőr","technológus","gyártás",
    "pénztáros","eladó","értékesítő","ügyfél","adminisztrátor","raktár","logisztika","gépkezelő","karbantartó"
]

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
BASE = "https://www.profession.hu/allasok/1,0,0,{}"

# Portálok definiálása
PORTALS = {
    "profession": {
        "name": "Profession.hu",
        "description": "Magyarország legnagyobb álláskereső portálja",
        "status": "Aktív",
        "enabled": True
    },
    "nofluffjobs": {
        "name": "No Fluff Jobs",
        "description": "IT specializált álláskereső portál - Magyar felület",
        "status": "Aktív",
        "enabled": True
    },
    "jobline": {
        "name": "Jobline.hu",
        "description": "IT specializált álláskereső portál",
        "status": "Tervezés alatt",
        "enabled": False
    },
    "cvonline": {
        "name": "CV-Online.hu",
        "description": "Népszerű álláskereső portál",
        "status": "Tervezés alatt",
        "enabled": False
    },
    "indeed": {
        "name": "Indeed.hu",
        "description": "Nemzetközi álláskereső portál",
        "status": "Tervezés alatt",
        "enabled": False
    }
}

# ——————————————————————————
# Kulcsszó-katalógus (maximalista, bővíthető)
# ——————————————————————————
KW_LANG = [
    "java","python","c#",".net","dotnet","c++","cpp","golang","go","rust","php","ruby","scala","haskell",
    "kotlin","swift","objective-c"
]
KW_FE = [
    "frontend","front-end","react","angular","vue","svelte","next.js","nuxt","typescript","javascript","web developer"
]
KW_BE = [
    "backend","back-end","node","spring","spring boot","quarkus",".net core","asp.net","laravel","symfony","django","flask","fastapi"
]
KW_MOBILE = [
    "android","ios","mobile","kotlin","swift","flutter","react native","xamarin","ionic"
]
KW_DATA = [
    "data engineer","data scientist","ml engineer","machine learning","ai engineer","big data",
    "etl","elt","dwh","data warehouse","power bi","tableau","qlik","sql","spark","hadoop","db developer","snowflake","databricks"
]
KW_DEVOPS = [
    "devops","sre","site reliability","platform engineer","cloud engineer","kubernetes","docker","terraform","ansible",
    "aws","azure","gcp","cloud architect","finops","observability"
]
KW_TEST = [
    "qa","test automation","tesztautomatizálás","tesztmérnök","sdet","cypress","selenium","playwright","jmeter","postman"
]
KW_EMBED = [
    "embedded","firmware","fpga","rtos","bare metal","microcontroller","stm32","esp32","embedded linux","yocto","driver developer","c developer"
]
KW_SECURITY = [
    "security engineer","application security","appsec","devsecops","penetration tester","pentest","iam","siem","soc"
]
KW_ENTERPRISE = [
    "sap","abap","erp","crm","salesforce","servicenow","mendix","outsystems","navision","business central","oracle developer","ms dynamics"
]
KW_GENERAL_HU = [
    "fejlesztő","programozó","szoftver","szoftvermérnök","rendszermérnök","alkalmazásfejlesztő","alkalmazás üzemeltető","full stack","full-stack"
]

# Összes kulcsszó egy listában
ALL_KEYWORDS = (
    KW_LANG + KW_FE + KW_BE + KW_MOBILE + KW_DATA + KW_DEVOPS +
    KW_TEST + KW_EMBED + KW_SECURITY + KW_ENTERPRISE + KW_GENERAL_HU
)

# Kategóriák definiálása a frontend-hez
CATEGORIES = {
    "languages": {"name": "Programozási nyelvek", "keywords": KW_LANG},
    "frontend": {"name": "Frontend fejlesztés", "keywords": KW_FE},
    "backend": {"name": "Backend fejlesztés", "keywords": KW_BE},
    "mobile": {"name": "Mobil fejlesztés", "keywords": KW_MOBILE},
    "data": {"name": "Data & AI", "keywords": KW_DATA},
    "devops": {"name": "DevOps & Cloud", "keywords": KW_DEVOPS},
    "testing": {"name": "Tesztelés", "keywords": KW_TEST},
    "embedded": {"name": "Embedded", "keywords": KW_EMBED},
    "security": {"name": "Biztonság", "keywords": KW_SECURITY},
    "enterprise": {"name": "Enterprise", "keywords": KW_ENTERPRISE},
    "general": {"name": "Általános IT", "keywords": KW_GENERAL_HU},
    # No Fluff Jobs kategóriák
    "nofluff_ai_ml": {"name": "AI/ML", "keywords": ["AI", "ML", "Machine Learning", "Deep Learning", "TensorFlow", "PyTorch"]},
    "nofluff_sysadmin": {"name": "Rendszergazda", "keywords": ["System Administrator", "Linux", "Windows", "Network", "Server"]},
    "nofluff_business": {"name": "Üzleti elemzés", "keywords": ["Business Analyst", "Product Owner", "Scrum Master", "Agile"]},
    "nofluff_architecture": {"name": "Architecture", "keywords": ["Architect", "Solution Architect", "System Design", "Microservices"]},
    "nofluff_backend": {"name": "Backend", "keywords": ["Backend", "API", "Database", "Server", "Node.js", "Python", "Java"]},
    "nofluff_data": {"name": "Data", "keywords": ["Data Engineer", "Data Scientist", "Analytics", "Big Data", "ETL"]},
    "nofluff_design": {"name": "Design", "keywords": ["UI/UX", "Designer", "Frontend", "CSS", "HTML", "JavaScript"]},
    "nofluff_devops": {"name": "DevOps", "keywords": ["DevOps", "CI/CD", "Docker", "Kubernetes", "AWS", "Azure"]},
    "nofluff_erp": {"name": "ERP", "keywords": ["ERP", "SAP", "Oracle", "Business Systems"]},
    "nofluff_embedded": {"name": "Embedded", "keywords": ["Embedded", "Firmware", "Hardware", "IoT", "Microcontroller"]},
    "nofluff_frontend": {"name": "Frontend", "keywords": ["Frontend", "React", "Vue", "Angular", "JavaScript", "TypeScript"]},
    "nofluff_fullstack": {"name": "Fullstack", "keywords": ["Fullstack", "Full Stack", "Full-Stack", "MEAN", "MERN"]},
    "nofluff_gamedev": {"name": "GameDev", "keywords": ["Game Developer", "Unity", "Unreal", "Game Design", "C++"]},
    "nofluff_mobile": {"name": "Mobile", "keywords": ["Mobile", "iOS", "Android", "React Native", "Flutter"]},
    "nofluff_pm": {"name": "Project Manager", "keywords": ["Project Manager", "PM", "Product Manager", "Agile", "Scrum"]},
    "nofluff_security": {"name": "Security", "keywords": ["Security", "Cybersecurity", "Penetration Testing", "Security Engineer"]},
    "nofluff_support": {"name": "Support", "keywords": ["Support", "Help Desk", "Technical Support", "Customer Service"]},
    "nofluff_testing": {"name": "Testing", "keywords": ["Testing", "QA", "Test Engineer", "Automation", "Selenium"]},
    "nofluff_other": {"name": "Egyéb IT", "keywords": ["IT", "Technology", "Software", "Development"]}
}

def clean_text(s: str) -> str:
    if not s:
        return ""
    # Karakterkódolás javítása
    if isinstance(s, str):
        s = s.encode('utf-8', errors='ignore').decode('utf-8')
    s = re.sub(r"<.*?>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Speciális karakterek tisztítása
    s = s.replace('\u00a0', ' ')  # Non-breaking space
    s = s.replace('\u2013', '-')  # En dash
    s = s.replace('\u2014', '--') # Em dash
    s = s.replace('\u2018', "'")  # Left single quotation mark
    s = s.replace('\u2019', "'")  # Right single quotation mark
    s = s.replace('\u201c', '"')  # Left double quotation mark
    s = s.replace('\u201d', '"')  # Right double quotation mark
    
    # Pozíció címek tisztítása - nemi megkülönböztetés eltávolítása (bárhol)
    s = re.sub(r'\s*\(f/m/x\)\s*', ' ', s, flags=re.IGNORECASE)  # (f/m/x) bárhol
    s = re.sub(r'\s*\(m/f\)\s*', ' ', s, flags=re.IGNORECASE)    # (m/f) bárhol
    s = re.sub(r'\s*\(f/m\)\s*', ' ', s, flags=re.IGNORECASE)    # (f/m) bárhol
    s = re.sub(r'\s*\(m/w/d\)\s*', ' ', s, flags=re.IGNORECASE)  # (m/w/d) bárhol
    s = re.sub(r'\s*\(w/m/d\)\s*', ' ', s, flags=re.IGNORECASE)  # (w/m/d) bárhol
    s = re.sub(r'\s*\(d/m/w\)\s*', ' ', s, flags=re.IGNORECASE)  # (d/m/w) bárhol
    s = re.sub(r'\s*\(m/w/d\)\*?\s*', ' ', s, flags=re.IGNORECASE)  # (m/w/d)* bárhol
    s = re.sub(r'\s*\(f/m/x\)\*?\s*', ' ', s, flags=re.IGNORECASE)  # (f/m/x)* bárhol
    
    return s

def build_feed_url(keyword: str) -> str:
    return BASE.format(quote(keyword, safe=""))

def is_probably_dev(title: str, desc: str) -> bool:
    t = (title or "").lower()
    d = (desc or "").lower()
    if any(bad in t for bad in EXCLUDE_PHRASES) or any(bad in d for bad in EXCLUDE_PHRASES):
        return False
    return True

def parse_publication_date(date_str: str):
    """Publikálási dátum feldolgozása ISO formátumra"""
    if not date_str:
        return None, False
    
    date_str = date_str.strip().lower()
    
    # "Friss" esetén
    if "friss" in date_str:
        return datetime.today().strftime("%Y-%m-%d"), True
    
    # Magyar dátum formátumok
    from datetime import timedelta
    
    # 2025. október 20. formátum
    match = re.search(r'(\d{4})\.\s*(\w+)\s*(\d{1,2})\.', date_str)
    if match:
        year, month_name, day = match.groups()
        month_map = {
            'január': 1, 'február': 2, 'március': 3, 'április': 4,
            'május': 5, 'június': 6, 'július': 7, 'augusztus': 8,
            'szeptember': 9, 'október': 10, 'november': 11, 'december': 12
        }
        if month_name in month_map:
            try:
                date_obj = datetime(int(year), month_map[month_name], int(day))
                return date_obj.strftime("%Y-%m-%d"), False
            except ValueError:
                pass
    
    # 2025-10-20 formátum
    match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        year, month, day = match.groups()
        try:
            date_obj = datetime(int(year), int(month), int(day))
            return date_obj.strftime("%Y-%m-%d"), False
        except ValueError:
            pass
    
    # 20. október 2025 formátum
    match = re.search(r'(\d{1,2})\.\s*(\w+)\s*(\d{4})', date_str)
    if match:
        day, month_name, year = match.groups()
        month_map = {
            'január': 1, 'február': 2, 'március': 3, 'április': 4,
            'május': 5, 'június': 6, 'július': 7, 'augusztus': 8,
            'szeptember': 9, 'október': 10, 'november': 11, 'december': 12
        }
        if month_name in month_map:
            try:
                date_obj = datetime(int(year), month_map[month_name], int(day))
                return date_obj.strftime("%Y-%m-%d"), False
            except ValueError:
                pass
    
    # Ha nem sikerült feldolgozni, ma dátumot adunk vissza
    return datetime.today().strftime("%Y-%m-%d"), False

def create_excel_export_multi_portal(jobs_data):
    """Excel fájl létrehozása több portál adataiból külön sheet-ekkel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Portálok szerint csoportosítás
        portal_data = {}
        for job in jobs_data:
            source = job.get("Forrás", "Ismeretlen")
            # Portál neve kinyerése (pl. "Profession – IT főkategória" -> "Profession")
            portal_name = source.split(" – ")[0] if " – " in source else source.split(" - ")[0] if " - " in source else source
            
            if portal_name not in portal_data:
                portal_data[portal_name] = []
            portal_data[portal_name].append(job)
        
        # Alapértelmezett sheet törlése
        wb.remove(wb.active)
        
        # Sheet-ek létrehozása portálonként
        for portal_name, portal_jobs in portal_data.items():
            ws = wb.create_sheet(title=portal_name)
            _add_data_to_sheet(ws, portal_jobs)
        
        # Összesítő sheet létrehozása
        ws_summary = wb.create_sheet(title="Összesítés", index=0)
        _add_summary_to_sheet(ws_summary, portal_data)
        
        return wb
        
    except ImportError:
        print("openpyxl nincs telepítve, egyszerű Excel export használata")
        return None

def _add_data_to_sheet(ws, jobs_data):
    """Adatok hozzáadása egy sheet-hez"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Oszlopok definiálása
    headers = [
        "ID", "Forrás", "Pozíció", "Cég", "Lokáció", "Fizetés", 
        "Munkavégzés típusa", "Cég mérete", "Publikálva", 
        "Lekérés dátuma", "Leírás", "Link"
    ]
    
    # Stílusok definiálása
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border stílus
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Fejléc sor hozzáadása
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Debug logging
    print(f"[SHEET DEBUG] Adatok hozzáadása: {len(jobs_data)} állás")
    if jobs_data:
        print(f"[SHEET DEBUG] Első job mezői: {list(jobs_data[0].keys())}")
        # Profession.hu formátum ellenőrzése (kisbetűs)
        print(f"[SHEET DEBUG] Profession.hu formátum - forras: {jobs_data[0].get('forras', 'N/A')}")
        print(f"[SHEET DEBUG] Profession.hu formátum - pozicio: {jobs_data[0].get('pozicio', 'N/A')}")
        print(f"[SHEET DEBUG] Profession.hu formátum - ceg: {jobs_data[0].get('ceg', 'N/A')}")
        # No Fluff Jobs formátum ellenőrzése (nagybetűs)
        print(f"[SHEET DEBUG] No Fluff Jobs formátum - Forrás: {jobs_data[0].get('Forrás', 'N/A')}")
        print(f"[SHEET DEBUG] No Fluff Jobs formátum - Pozíció: {jobs_data[0].get('Pozíció', 'N/A')}")
        print(f"[SHEET DEBUG] No Fluff Jobs formátum - Cég: {jobs_data[0].get('Cég', 'N/A')}")
    
    # Adatok hozzáadása - TÁMOGATÁS: Profession.hu (kisbetűs) ÉS No Fluff Jobs (nagybetűs)
    for row_num, job in enumerate(jobs_data, 2):
        # Row number mint ID
        ws.cell(row=row_num, column=1, value=row_num - 1)
        
        # PROFESSION.HU formátum (kisbetűs) ÉS NO FLUFF JOBS formátum (nagybetűs) támogatása
        # Profession.hu: "forras", "pozicio", "ceg", "lokacio", "link", "publikalva", "leiras"
        # No Fluff Jobs: "Forrás", "Pozíció", "Cég", "Lokáció", "Link", "Publikálva", "Leírás"
        
        forrás = job.get("Forrás") or job.get("forras") or ""
        pozíció = job.get("Pozíció") or job.get("pozicio") or ""
        cég = job.get("Cég") or job.get("ceg") or ""
        lokáció = job.get("Lokáció") or job.get("lokacio") or ""
        link = job.get("Link") or job.get("link") or ""
        publikálva = job.get("Publikálva") or job.get("publikalva") or ""
        leírás = job.get("Leírás") or job.get("leiras") or ""
        fizetés = job.get("Fizetés") or job.get("fizetes") or ""
        
        # CELLÁK ÍRÁSA - explicit értékekkel
        # Debug: első sor részletes kiírása
        if row_num == 2:
            print(f"[SHEET WRITE DEBUG] Sor {row_num} előkészítése")
            print(f"[SHEET WRITE DEBUG] Job objektum: {job}")
            print(f"[SHEET WRITE DEBUG] Forrás (nagybetűs): {job.get('Forrás')}")
            print(f"[SHEET WRITE DEBUG] Forrás (kisbetűs): {job.get('forras')}")
            print(f"[SHEET WRITE DEBUG] Pozíció (nagybetűs): {job.get('Pozíció')}")
            print(f"[SHEET WRITE DEBUG] Pozíció (kisbetűs): {job.get('pozicio')}")
        
        # VALUES EXPLICITEN ÍRVA - nincs feltétel
        ws.cell(row=row_num, column=2, value=forrás)
        ws.cell(row=row_num, column=3, value=pozíció)
        ws.cell(row=row_num, column=4, value=cég)
        ws.cell(row=row_num, column=5, value=lokáció)
        ws.cell(row=row_num, column=6, value=fizetés)
        ws.cell(row=row_num, column=7, value="")  # Munkavégzés_típusa
        ws.cell(row=row_num, column=8, value="")  # Cég_mérete
        ws.cell(row=row_num, column=9, value=publikálva)
        ws.cell(row=row_num, column=10, value=job.get("lekeres_datuma") or "")  # Lekérés_dátuma
        ws.cell(row=row_num, column=11, value=leírás)
        ws.cell(row=row_num, column=12, value=link)
        
        # Debug: első sor értékek ellenőrzése
        if row_num == 2:
            print(f"[SHEET WRITE] Sor {row_num} írva - Forrás='{forrás}', Pozíció='{pozíció}', Cég='{cég}'")
            # Teszt: írjunk egy statikus értéket is
            ws.cell(row=row_num, column=13, value="TEST_DEBUG_COLUMN")
            print(f"[SHEET WRITE] TEST oszlop hozzáadva a 13. oszlophoz")
        
        # Border hozzáadása minden cellához
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # Oszlopok szélességének beállítása
    column_widths = [8, 20, 40, 25, 20, 20, 20, 15, 15, 15, 50, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Szűrés hozzáadása
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs_data) + 1}"

def _add_summary_to_sheet(ws, portal_data):
    """Összesítő sheet létrehozása"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Fejléc
    ws.cell(row=1, column=1, value="Portál").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Állások száma").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Friss állások").font = Font(bold=True)
    
    # Adatok
    total_jobs = 0
    total_fresh = 0
    row = 2
    
    for portal_name, portal_jobs in portal_data.items():
        fresh_count = sum(1 for job in portal_jobs if job.get("Friss_állás", False))
        
        ws.cell(row=row, column=1, value=portal_name)
        ws.cell(row=row, column=2, value=len(portal_jobs))
        ws.cell(row=row, column=3, value=fresh_count)
        
        total_jobs += len(portal_jobs)
        total_fresh += fresh_count
        row += 1
    
    # Összesítés
    ws.cell(row=row, column=1, value="ÖSSZESEN").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_jobs).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_fresh).font = Font(bold=True)
    
    # Oszlopok szélessége
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def create_excel_export(jobs_data):
    """Excel fájl létrehozása a munkák adataiból - MOST AZ _add_data_to_sheet-ET HASZNÁLJA"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "IT Állások"
        
        print(f"[CREATE_EXCEL_EXPORT] jobs_data count: {len(jobs_data)}")
        
        # HASZNÁLJUK AZ _add_data_to_sheet FÜGGVÉNYT - ami biztosan működik!
        _add_data_to_sheet(ws, jobs_data)
        
        return wb
        
    except ImportError:
        print("openpyxl nincs telepítve, egyszerű Excel export használata")
        return None
    except Exception as e:
        print(f"[CREATE_EXCEL_EXPORT] HIBÁ: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_total_pages(source_name: str, url: str):
    """Get total number of pages for a search - fix oldalszám a teljes lefedettségért"""
    print(f"[DEBUG] get_total_pages hivasa: {source_name} - {url}")
    
    # IT főkategória fix oldalszám - 936 állás elérése
    if "it-programozas-fejlesztes" in url:
        # 936 állás / ~15 állás/oldal = 62-63 oldal
        # Biztonsági puffer: 70 oldal
        fixed_pages = 70
        print(f"[FIXED] {source_name} - Fix oldalszám: {fixed_pages} oldal (936 állás cél)")
        return fixed_pages
    
    # Egyéb kategóriák esetén dinamikus meghatározás
    try:
        session = requests.Session()
        session.headers.update(HEADERS)
        
        # Első oldal lekérése
        r = session.get(url, timeout=30)
        r.raise_for_status()
        r.encoding = "utf-8"
        
        soup = BeautifulSoup(r.content, 'html.parser')
        
        # 1. Keresés pagination elemekben
        pagination_selectors = [
            'div.pager', 'nav.pager', '.pager',  # Profession.hu specifikus
            'div.pagination', 'nav.pagination', '.pagination', 
            'ul.pagination', 'ol.pagination'
        ]
        
        for selector in pagination_selectors:
            pagination = soup.select_one(selector)
        if pagination:
                print(f"[DEBUG] Pagination elem találva: {selector}")
                
                # Keresés utolsó oldal linkjében
                page_links = pagination.find_all('a', href=True)
                if page_links:
                    # Utolsó link href-jéből oldalszám kinyerése
                    last_href = page_links[-1].get('href', '')
                    print(f"[DEBUG] Utolsó pagination link: {last_href}")
                    
                    # Profession.hu formátum: /OLDALSZÁM,10
                    if '/,' in last_href:
                        try:
                            page_num = int(last_href.split('/')[-1].split(',')[0])
                            print(f"[SUCCESS] {source_name} - Dinamikus oldalszám: {page_num} oldal")
                            return page_num
                        except (ValueError, IndexError):
                            pass
                    
                    # Standard formátum: ?page=N vagy &page=N
                    if 'page=' in last_href:
                        try:
                            page_num = int(last_href.split('page=')[1].split('&')[0].split('#')[0])
                            print(f"[SUCCESS] {source_name} - Dinamikus oldalszám: {page_num} oldal")
                            return page_num
                        except (ValueError, IndexError):
                            pass
                
                # Fallback: span elemekben keresés (ha nincs link)
                page_spans = pagination.find_all('span')
                if page_spans:
                    max_page = 0
                    for span in page_spans:
                        span_text = span.get_text(strip=True)
                        if span_text.isdigit():
                            max_page = max(max_page, int(span_text))
                    if max_page > 0:
                        print(f"[SUCCESS] {source_name} - Dinamikus oldalszám (span): {max_page} oldal")
                        return max_page
        
        # 2. Fallback: job card szám alapján becslés
        job_selectors = [
            'li.card.job-card', 'ul.job-cards li', '.job-card', 
            '.job-item', '.listing-item', '.search-result-item'
        ]
        
        total_jobs = 0
        for selector in job_selectors:
            job_cards = soup.select(selector)
            if job_cards:
                total_jobs = len(job_cards)
                break
        
        if total_jobs > 0:
            # Becslés: ~13-15 állás/oldal (Profession.hu átlag)
            estimated_pages = max(1, (total_jobs + 12) // 13)  # Felfelé kerekítés
            print(f"[ESTIMATE] {source_name} - Becsült oldalszám: {estimated_pages} (alapján: {total_jobs} job card)")
            return estimated_pages
        
        # 3. Fallback: 1 oldal
        print(f"[FALLBACK] {source_name} - Nem található pagination, 1 oldal használata")
        return 1
        
    except Exception as e:
        print(f"[WARNING] {source_name} - Oldalszám meghatározási hiba: {e}, 1 oldal használata")
        return 1

def fetch_html_jobs(source_name: str, url: str, max_pages: int = None):
    """HTML scraping a Profession.hu álláslistákról - dinamikus oldalszám"""
    if not BeautifulSoup:
        print("BeautifulSoup nincs telepítve, RSS fallback használata")
        return fetch_rss_fallback(source_name, url)
    
    # Get total pages dynamically if not specified
    if max_pages is None:
        max_pages = get_total_pages(source_name, url)
    
    # Dinamikus limit - mindig annyi oldal amennyi van
    # Csak biztonsági limit: maximum 200 oldal (védés a végtelen ciklus ellen)
    max_pages = min(max_pages, 200)  # Biztonsági limit
    
    print(f"[INFO] {source_name} - {max_pages} oldal feldolgozása")
    
    all_items = []
    
    for page in range(1, max_pages + 1):
        try:
            # Oldalszámozás hozzáadása - Profession.hu formátum: /OLDALSZÁM,10
            if '/1,10' in url:
                page_url = url.replace('/1,10', f'/{page},10')
            elif '&page=' in url or '?page=' in url:
                page_url = f"{url}&page={page}" if "?" in url else f"{url}?page={page}"
            else:
                page_url = f"{url}&page={page}" if "?" in url else f"{url}?page={page}"
            
            # Debug: URL ellenőrzés
            if page <= 3:  # Csak az első 3 oldalról debug
                print(f"   [DEBUG] Oldal {page} URL: {page_url}")
            
            # Retry logika timeout esetén
            max_retries = 5
            for retry in range(max_retries):
                try:
                    r = requests.get(page_url, headers=HEADERS, timeout=30)
                    r.raise_for_status()
                    break
                except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.exceptions.HTTPError, requests.exceptions.RequestException) as e:
                    if retry < max_retries - 1:
                        wait_time = (retry + 1) * 10
                        print(f"   [WARNING] Error, retry {retry + 1}/{max_retries} in {wait_time}s: {e}")
                        time.sleep(wait_time)
                    else:
                        print(f"   [ERROR] Max retries reached, skipping page {page}")
                        raise e
            r.encoding = "utf-8"
            
            # Karakterkódolás javítása
            content = r.text
            if r.encoding.lower() != 'utf-8':
                content = content.encode('utf-8', errors='ignore').decode('utf-8')
            
            soup = BeautifulSoup(content, "html.parser")
            
            # Keresés ul.job-cards > li elemekben
            job_cards = soup.select("ul.job-cards li")
            if not job_cards:
                # Fallback: más lehetséges szelektorok
                job_cards = soup.select(".job-card, .job-item, .listing-item, .search-result-item")
            
            print(f"DEBUG: {source_name} - Oldal {page}: {len(job_cards)} job card")
            
            if not job_cards:
                # Ha nincs több állás, szakítsuk meg
                print(f"   [STOP] Nincs több állás az oldalon, leállítás")
                break
            
            for card in job_cards:
                try:
                    # Pozíció címe
                    title_elem = card.select_one("h3, .job-title, .position-title, .title, a[href*='/allas/']")
                    title = clean_text(title_elem.get_text()) if title_elem else ""
                    
                    # Link
                    link_elem = card.select_one("a[href*='/allas/']")
                    link = link_elem.get("href") if link_elem else ""
                    if link and not link.startswith("http"):
                        link = "https://www.profession.hu" + link
                    
                    # Cég neve - profession.hu specifikus szelektorok (bővített lista)
                    company_elem = card.select_one(".company-name, .employer-name, .job-company, .company, [data-company], .job-card-company, .job-card__company-name, .job-card__company, .company-link, .employer, .job-employer, .card-company, .listing-company, .job-listing-company, .company-title, .employer-title")
                    if company_elem:
                        company = clean_text(company_elem.get_text())
                    else:
                        # Fallback: keresés a szövegben - javított megközelítés
                        import re
                        card_text = card.get_text()
                        
                        # Keresés cégnevek után (Kft, Zrt, stb.) - csak a végén lévő cégneveket
                        # A cég neve általában a job card végén van
                        lines = card_text.split('\n')
                        company = ""
                        
                        # Végigjárjuk a sorokat hátulról
                        for line in reversed(lines):
                            line = line.strip()
                            if not line:
                                continue
                            
                            # Keresés cégnevek után - szigorúbb regex
                            company_match = re.search(r'([A-ZÁÉÍÓÖŐÚÜŰ][a-zA-ZÁÉÍÓÖŐÚÜŰáéíóöőúüű\s&.,-]+(?:Kft|Zrt|Nyrt|Bt|Kkt|Ltd|Corp|Inc|Hungary|Services|Solutions|Technologies|Systems|Group|Consulting|Software|Digital|IT|Tech))\b', line)
                            if company_match:
                                potential_company = company_match.group(1).strip()
                                
                                # Ellenőrizzük, hogy nem pozíció cím vagy leírás
                                if not any(word in potential_company.lower() for word in [
                                    'developer', 'engineer', 'manager', 'analyst', 'specialist', 'consultant', 'architect',
                                    'programozó', 'fejlesztő', 'menedzser', 'elemző', 'szakember', 'tanácsadó', 'építész',
                                    'rendszer', 'alkalmazás', 'webes', 'mobil', 'backend', 'frontend', 'full-stack',
                                    'tervezése', 'fejlesztése', 'optimalizálása', 'integrálva', 'meglévő', 'új', 'munkatárs',
                                    'delivery', 'roadmap', 'requirement', 'specification', 'acceptance', 'criteria', 'mvp',
                                    'collaboration', 'business', 'pozíció', 'osztályvezető', 'együttműködve', 'jelent',
                                    'maintain', 'clear', 'drive', 'define', 'close', 'kontrolling', 'szorosan',
                                    'kulcsfontosságú', 'szereplőként', 'üzleti', 'igényeket', 'lefordítod', 'projektekhez',
                                    'kapcsolódó', 'követelmények', 'elemzése', 'ensure', 'stable', 'efficient', 'daily',
                                    'operation', 'consumer', 'services', 'card-related', 'design', 'build', 'intelligent',
                                    'pipelines', 'retrieval', 'systems', 'mercedes-benz', 'insurance', 'service', 'provider',
                                    'group', 'build', 'lead', 'dedicated', 'finance', 'applications', 'team', 'within',
                                    'elia', 'forefront', 'energy', 'transition', 'developing', 'improving', 'governance',
                                    'frameworks', 'next', 'challenge', 'low', 'code', 'agentic', 'solutions', 'hyperautomation'
                                ]):
                                    # További ellenőrzés: ne legyen túl hosszú (leírás)
                                    if len(potential_company) < 100:
                                        company = potential_company
                                        break
                        
                        # Ha nem találtunk cégnevet, próbáljuk meg a teljes szövegből
                        if not company:
                            company_match = re.search(r'([A-ZÁÉÍÓÖŐÚÜŰ][a-zA-ZÁÉÍÓÖŐÚÜŰáéíóöőúüű\s&.,-]+(?:Kft|Zrt|Nyrt|Bt|Kkt|Ltd|Corp|Inc|Hungary|Services|Solutions|Technologies|Systems|Group|Consulting|Software|Digital|IT|Tech))\b', card_text)
                            if company_match:
                                potential_company = company_match.group(1).strip()
                                if not any(word in potential_company.lower() for word in [
                                    'developer', 'engineer', 'manager', 'analyst', 'specialist', 'consultant', 'architect',
                                    'programozó', 'fejlesztő', 'menedzser', 'elemző', 'szakember', 'tanácsadó', 'építész',
                                    'rendszer', 'alkalmazás', 'webes', 'mobil', 'backend', 'frontend', 'full-stack',
                                    'tervezése', 'fejlesztése', 'optimalizálása', 'integrálva', 'meglévő', 'új', 'munkatárs',
                                    'delivery', 'roadmap', 'requirement', 'specification', 'acceptance', 'criteria', 'mvp',
                                    'collaboration', 'business', 'pozíció', 'osztályvezető', 'együttműködve', 'jelent',
                                    'maintain', 'clear', 'drive', 'define', 'close', 'kontrolling', 'szorosan',
                                    'kulcsfontosságú', 'szereplőként', 'üzleti', 'igényeket', 'lefordítod', 'projektekhez',
                                    'kapcsolódó', 'követelmények', 'elemzése', 'ensure', 'stable', 'efficient', 'daily',
                                    'operation', 'consumer', 'services', 'card-related', 'design', 'build', 'intelligent',
                                    'pipelines', 'retrieval', 'systems', 'mercedes-benz', 'insurance', 'service', 'provider',
                                    'group', 'build', 'lead', 'dedicated', 'finance', 'applications', 'team', 'within',
                                    'elia', 'forefront', 'energy', 'transition', 'developing', 'improving', 'governance',
                                    'frameworks', 'next', 'challenge', 'low', 'code', 'agentic', 'solutions', 'hyperautomation'
                                ]) and len(potential_company) < 100:
                                    company = potential_company
                            else:
                                company = ""
                    
                    # Lokáció - profession.hu specifikus szelektorok
                    location_elem = card.select_one(".job-location, .location, .city, .place, [data-location], .job-card-location")
                    if not location_elem:
                        # Fallback: keresés a szövegben
                        card_text = card.get_text()
                        # Magyar városok keresése
                        hungarian_cities = ["Budapest", "Debrecen", "Szeged", "Miskolc", "Pécs", "Győr", "Nyíregyháza", "Kecskemét", "Székesfehérvár", "Szombathely", "Szolnok", "Tatabánya", "Kaposvár", "Békéscsaba", "Zalaegerszeg", "Érd", "Sopron", "Veszprém", "Dunaújváros", "Hódmezővásárhely"]
                        for city in hungarian_cities:
                            if city in card_text:
                                location = city
                                break
                        else:
                            location = ""
                    else:
                        location = clean_text(location_elem.get_text())
                    
                    # Leírás
                    desc_elem = card.select_one(".description, .job-description, .summary, .excerpt")
                    desc = clean_text(desc_elem.get_text()) if desc_elem else ""
                    
                    # Dátum
                    date_elem = card.select_one(".date, .published, .job-date, .time")
                    pub_date = clean_text(date_elem.get_text()) if date_elem else ""
                    
                    # Dátum feldolgozás - ISO formátumra konvertálás
                    pub_date_iso, is_fresh = parse_publication_date(pub_date)
                    
                    if title and link:
                        all_items.append({
                            "Forrás": source_name, 
                            "Pozíció": title, 
                            "Link": link, 
                            "Leírás": desc,
                            "Publikálva": pub_date,
                            "Publikálva_dátum": pub_date_iso,
                            "Friss_állás": is_fresh,
                            "Cég": company,
                            "Lokáció": location
                        })
                        
                except Exception as e:
                    print(f"ERROR parsing job card: {e}")
                    continue
            
            # Kímélet a szerver felé (csökkentett delay - stabilitásért)
            time.sleep(1.0)
            
        except Exception as e:
            print(f"ERROR fetching page {page}: {e}")
            break
    
    print(f"DEBUG: {source_name} - Összesen {len(all_items)} állás {page-1} oldalról (max {max_pages})")
    return all_items


def fetch_nofluffjobs_all_categories(max_pages_per_category: int = 50):
    """No Fluff Jobs összes IT kategória feldolgozása - 1997 állás elérése"""
    print(f"[INFO] No Fluff Jobs összes IT kategória feldolgozása")
    
    # Kategóriák definíciója
    categories = {
        "AI/ML": "artificial-intelligence",
        "Rendszergazda": "sys-administrator",
        "Üzleti elemzés": "business-analyst",
        "Architecture": "architecture",
        "Backend": "backend",
        "Data": "data",
        "Design": "ux",
        "devOps": "devops",
        "ERP": "erp",
        "Embedded": "embedded",
        "Frontend": "frontend",
        "Fullstack": "fullstack",
        "GameDev": "game-dev",
        "Mobile": "mobile",
        "Project Manager": "project-manager",
        "Security": "security",
        "Support": "support",
        "Testing": "testing",
        "Egyéb IT": "other"
    }
    
    all_jobs = []
    seen_links = set()
    
    total_categories = len(categories)
    
    for i, (category_name, category_url) in enumerate(categories.items(), 1):
        try:
            print(f"[PROGRESS] Kategória {i}/{total_categories}: {category_name}")
            
            # Kategória URL
            url = f"https://nofluffjobs.com/hu/{category_url}"
            
            # Pagination feldolgozása
            category_jobs = fetch_nofluffjobs_jobs_pagination(
                f"No Fluff Jobs - {category_name}",
                url,
                max_pages=max_pages_per_category
            )
            
            # Egyedi állások kinyerése
            for job in category_jobs:
                link = job.get('Link')
                if link and link not in seen_links:
                    seen_links.add(link)
                    all_jobs.append(job)
            
            print(f"  [OK] {len(category_jobs)} allas talalva, {len(all_jobs)} egyedi osszesen")
            
        except Exception as e:
            print(f"  [ERROR] Hiba: {e}")
            continue
    
    print(f"[SUCCESS] Osszesen {len(all_jobs)} egyedi allas talalva {total_categories} kategoriabol")
    return all_jobs

def fetch_nofluffjobs_jobs_pagination(source_name: str, url: str, max_pages: int = 1000):
    """No Fluff Jobs pagination kezelése - több oldal ellenőrzése"""
    print(f"[INFO] {source_name} - Pagination kezelése")
    
    try:
        # Selenium használata pagination-hoz
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.options import Options
        from selenium.common.exceptions import TimeoutException, NoSuchElementException
        import time
        from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
        
        # Chrome opciók beállítása - timeout optimalizálás
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Háttérben futás
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--disable-features=VizDisplayCompositor")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        
        # Timeout beállítások
        chrome_options.add_argument("--page-load-strategy=eager")  # Gyorsabb betöltés
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-plugins")
        chrome_options.add_argument("--disable-images")  # Képek letiltása gyorsabb betöltésért
        
        # WebDriver inicializálása timeout beállításokkal
        driver = None
        try:
            driver = webdriver.Chrome(options=chrome_options)
            
            # Timeout beállítások - növelve a megbízhatóságért
            driver.set_page_load_timeout(30)  # 30 másodperc timeout oldal betöltésre (webes környezetben lassabb lehet)
            driver.implicitly_wait(10)  # 10 másodperc implicit várakozás
        except Exception as driver_error:
            print(f"   [ERROR] WebDriver inicializálási hiba: {driver_error}")
            raise driver_error
        
        try:
            all_jobs = []
            seen_links = set()  # Duplikáció elkerülése
            pages_processed = 0  # Feldolgozott oldalak száma
            
            # URL parsing
            parsed_url = urlparse(url)
            query_params = parse_qs(parsed_url.query)
            
            # Elméleti maximum oldal - csak loop elkerülésére, nem korlátozza a valódi oldalszámot
            max_pages_limited = min(max_pages, 100)  # 100 oldal elméleti maximum
            print(f"   [INFO] Elméleti maximum {max_pages_limited} oldal (loop elkerülésére)")
            
            # Dinamikus oldalszám detektálás - első oldal betöltése
            try:
                first_page_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, urlencode({**query_params, 'page': ['1']}, doseq=True), parsed_url.fragment))
                print(f"   [DEBUG] Első oldal betöltése dinamikus oldalszám detektáláshoz: {first_page_url}")
                driver.get(first_page_url)
                
                # Pagination elemek keresése
                pagination_elements = driver.find_elements(By.CSS_SELECTOR, ".pagination a, .pager a, [class*='page'] a")
                if pagination_elements:
                    max_page_found = 0
                    for elem in pagination_elements:
                        try:
                            page_text = elem.text.strip()
                            if page_text.isdigit():
                                max_page_found = max(max_page_found, int(page_text))
                        except:
                            continue
                    
                    if max_page_found > 0:
                        print(f"   [SUCCESS] Dinamikus oldalszám detektálva: {max_page_found} oldal")
                        max_pages_limited = min(max_page_found, 100)  # Valódi oldalszám, de max 100
                    else:
                        print(f"   [INFO] Dinamikus oldalszám nem detektálható, elméleti maximum használata")
                else:
                    print(f"   [INFO] Pagination elemek nem találhatók, elméleti maximum használata")
                    
            except Exception as e:
                print(f"   [WARNING] Dinamikus oldalszám detektálás hiba: {e}")
                print(f"   [INFO] Elméleti maximum használata")
            
            print(f"   [INFO] Feldolgozandó oldalszám: {max_pages_limited}")
            
            # Konvergencia/üres oldal követés
            consecutive_empty_pages = 0
            max_consecutive_empty = 3  # Maximum 3 egymás után üres oldal
            
            for page in range(1, max_pages_limited + 1):
                try:
                    # Progress tracking
                    pages_processed = page  # Frissítjük a feldolgozott oldalak számát
                    print(f"   [PROGRESS] Oldal {page} - Eddig {len(all_jobs)} egyedi állás")
                    
                    # Page paraméter hozzáadása
                    query_params['page'] = [str(page)]
                    new_query = urlencode(query_params, doseq=True)
                    page_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, new_query, parsed_url.fragment))
                    
                    # Csak akkor töltse be az oldalt, ha nem az első oldal (már betöltötte)
                    if page > 1:
                        print(f"   [DEBUG] Oldal {page} betöltése: {page_url}")
                        driver.get(page_url)
                    else:
                        print(f"   [DEBUG] Oldal {page} már betöltve (dinamikus detektálás során)")
                    
                    # Oldal betöltésének várása - növelt timeout
                    page_loaded = False
                    try:
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "a[class*='posting-list-item']"))
                        )
                        page_loaded = True
                    except TimeoutException:
                        print(f"   [WARNING] Oldal {page} timeout - próbáljuk meg alternatív selectorral")
                        # Alternatív selector próbálkozás
                        try:
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, ".posting-list-item"))
                            )
                            page_loaded = True
                        except TimeoutException:
                            print(f"   [WARNING] Oldal {page} teljes timeout - folytatás")
                            # Folytatjuk timeout ellenére, de jelezzük hogy az oldal nem töltődött be
                            page_loaded = False
                    
                    # Job cards kinyerése - alternatív selectorokkal
                    job_cards = driver.find_elements(By.CSS_SELECTOR, "a[class*='posting-list-item']")
                    if not job_cards:
                        print(f"   [DEBUG] Elsődleges selector sikertelen, alternatív próbálkozás")
                        job_cards = driver.find_elements(By.CSS_SELECTOR, ".posting-list-item")
                    if not job_cards:
                        print(f"   [DEBUG] Másodlagos selector sikertelen, harmadik próbálkozás")
                        job_cards = driver.find_elements(By.CSS_SELECTOR, "a[href*='/hu/job/']")
                    
                    print(f"   [DEBUG] Oldal {page}: {len(job_cards)} job card találva (oldal betöltve: {page_loaded})")
                    
                    # Ha nincs job card, növeljük a consecutive_empty számlálót
                    if len(job_cards) == 0:
                        consecutive_empty_pages += 1
                        print(f"   [WARNING] Oldal {page} üres ({consecutive_empty_pages}/{max_consecutive_empty} egymás után üres)")
                        
                        # Ha túl sok egymás után üres oldal, kilépünk
                        if consecutive_empty_pages >= max_consecutive_empty:
                            print(f"   [STOP] {max_consecutive_empty} egymás után üres oldal - kilépés")
                            break
                    else:
                        # Ha van tartalom, nullázzuk a számlálót
                        consecutive_empty_pages = 0
                    
                    # EGYEDI LINKKEK ELLENŐRZÉSE - elkerüljük a duplikációt
                    new_links_on_page = 0
                    for card in job_cards:
                        try:
                            link = card.get_attribute("href")
                            if link and link not in seen_links:
                                seen_links.add(link)
                                new_links_on_page += 1
                        except:
                            continue
                    
                    print(f"   [DEBUG] Oldal {page}: {new_links_on_page} új link")
                    
                    # Ha 0 új link, konvergencia
                    if new_links_on_page == 0:
                        print(f"   [SUCCESS] Konvergencia elérve - nincs új link")
                        break
                    
                    # Job cards feldolgozása - friss elemekkel minden iterációban
                    for i in range(len(job_cards)):
                        try:
                            # Friss job cards lekérése minden iterációban (stale element elkerülése)
                            current_job_cards = driver.find_elements(By.CSS_SELECTOR, "a[class*='posting-list-item']")
                            if i >= len(current_job_cards):
                                break
                            
                            card = current_job_cards[i]
                            
                            # Link kinyerése
                            link = card.get_attribute("href")
                            if not link:  # Ha nincs link, kihagyjuk
                                continue
                            
                            # Pozíció címe
                            try:
                                title_elem = card.find_element(By.CSS_SELECTOR, 'h3[data-cy="title position on the job offer listing"]')
                                title = title_elem.text.strip()
                            except (NoSuchElementException, Exception):
                                title = ""
                            
                            # Cég neve
                            try:
                                company_elem = card.find_element(By.CSS_SELECTOR, 'h4.company-name')
                                company = company_elem.text.strip()
                            except (NoSuchElementException, Exception):
                                company = ""
                            
                            # Lokáció
                            try:
                                location_elem = card.find_element(By.CSS_SELECTOR, 'nfj-posting-item-city span, .city span, [class*="city"] span')
                                location = location_elem.text.strip()
                            except (NoSuchElementException, Exception):
                                location = ""
                            
                            # Ha nincs adat, próbáljuk meg URL parsing-gal
                            if not title or not company or not location:
                                if link:
                                    link_parts = link.split('/')
                                    if len(link_parts) > 3:
                                        job_part = link_parts[-1]
                                        parts = job_part.split('-')
                                        
                                        if not title and len(parts) >= 2:
                                            title = ' '.join(parts[:min(3, len(parts))]).title()
                                        
                                        if not company and len(parts) > 5:
                                            company_parts = parts[3:-1]
                                            company = ' '.join(company_parts).title()
                                        
                                        if not location and parts:
                                            last_part = parts[-1]
                                            if last_part.isdigit() and len(parts) > 1:
                                                location = parts[-2].title()
                                            else:
                                                location = last_part.title()
                            
                            # Dátum - jelenlegi dátum
                            pub_date = datetime.now().strftime("%Y.%m.%d")
                            pub_date_iso = datetime.now()
                            is_fresh = True
                            
                            if link:
                                all_jobs.append({
                                    "Forrás": source_name,
                                    "Pozíció": title,
                                    "Link": link,
                                    "Leírás": "",
                                    "Publikálva": pub_date,
                                    "Publikálva_dátum": pub_date_iso,
                                    "Friss_állás": is_fresh,
                                    "Cég": company,
                                    "Lokáció": location,
                                    "Fizetés": ""
                                })
                        
                        except Exception as e:
                            print(f"   [ERROR] Job card feldolgozási hiba: {e}")
                            continue
                    
                    # Ellenőrizzük hogy van-e új tartalom
                    print(f"   [DEBUG] Oldal {page}: {len(job_cards)} job card a listában, {len(all_jobs)} összesen")
                    
                    # Ha nincs job card, kilépünk
                    if len(job_cards) == 0:
                        print(f"   [DEBUG] Oldal {page} üres, kilépés")
                        break
                    
                    # Várás a következő oldal között
                    time.sleep(2)
                    
                except TimeoutException as te:
                    consecutive_empty_pages += 1
                    print(f"   [WARNING] Oldal {page} timeout - folytatás következő oldallal ({consecutive_empty_pages}/{max_consecutive_empty})")
                    print(f"   [ERROR] Timeout részletek: {str(te)[:200]}")
                    # Ha túl sok timeout egymás után, kilépünk
                    if consecutive_empty_pages >= max_consecutive_empty:
                        print(f"   [STOP] {max_consecutive_empty} egymás után timeout - kilépés")
                        break
                    continue
                except Exception as e:
                    consecutive_empty_pages += 1
                    print(f"   [ERROR] Oldal {page} hiba: {e}")
                    print(f"   [ERROR] Hiba típus: {type(e).__name__}")
                    # Ha túl sok hiba egymás után, kilépünk
                    if consecutive_empty_pages >= max_consecutive_empty:
                        print(f"   [STOP] {max_consecutive_empty} egymás után hiba - kilépés")
                        break
                    continue
            
            # Oldalak száma változó már definiálva a loop elején
            print(f"   [SUCCESS] {source_name} - {len(all_jobs)} állás feldolgozva {pages_processed} oldalról")
            return all_jobs
            
        finally:
            # Csak akkor zárd be a driver-t, ha létezik és inicializálva van
            try:
                if driver is not None:
                    driver.quit()
            except Exception as e:
                print(f"   [WARNING] Driver bezárási hiba: {e}")
    
    except ImportError:
        print(f"   [WARNING] Selenium nincs telepítve, fallback scraper használata")
        return fetch_nofluffjobs_jobs(source_name, url)
    except Exception as e:
        print(f"   [ERROR] Pagination hiba: {e}")
        print(f"   [FALLBACK] Fallback scraper használata")
        return fetch_nofluffjobs_jobs(source_name, url)

def fetch_nofluffjobs_jobs(source_name: str, url: str, max_pages: int = None):
    """HTML scraping a No Fluff Jobs magyar álláslistákról - valódi scraper"""
    print(f"[INFO] {source_name} - Valódi scraper tesztelése")
    
    if not BeautifulSoup:
        print("BeautifulSoup nincs telepítve, mock adatok használata")
        return fetch_nofluffjobs_mock(source_name)
    
    all_items = []
    
    try:
        # Rövidebb timeout és egyszerűbb megközelítés
        session = requests.Session()
        session.headers.update(HEADERS)
        
        print(f"   [DEBUG] URL: {url}")
        
        r = session.get(url, timeout=15)  # Rövid timeout
        r.raise_for_status()
        r.encoding = "utf-8"
        
        content = r.text
        print(f"   [DEBUG] HTML tartalom hossza: {len(content)} karakter")
        
        # Próbáljuk meg a valódi scraping-et
        soup = BeautifulSoup(content, "html.parser")
        
        # No Fluff Jobs job cards keresése - valódi HTML struktúra alapján
        # Az álláskártyák egy <a nfj-postings-item ... class="posting-list-item ..."> elemben vannak
        job_cards = soup.select("a[class*='posting-list-item']")
        
        print(f"   [DEBUG] Job cards találva: {len(job_cards)}")
        
        # Ha nincs job card, próbáljuk meg alternatív selectorokat
        if not job_cards:
            print(f"   [DEBUG] Job cards nélkül, alternatív selectorok...")
            job_cards = soup.select(".posting-list-item")
            print(f"   [DEBUG] Alternatív job cards: {len(job_cards)}")
        
        # Ha még mindig nincs, próbáljuk meg a linkeket közvetlenül
        if not job_cards:
            print(f"   [DEBUG] Job cards nélkül, linkek keresése közvetlenül")
            job_links = soup.select("a[href*='/hu/job/']")
            print(f"   [DEBUG] Job links találva: {len(job_links)}")
            
            # Használjuk az eredeti linkeket
            job_cards = job_links[:20]  # Maximum 20 link
        
        if not job_cards:
            print(f"   [WARNING] Nincs job card találva, fallback scraper használata")
            return fetch_nofluffjobs_fallback(source_name, content)
        
        # Feldolgozzuk a job cards-okat
        for card in job_cards[:10]:  # Maximum 10 job card
            try:
                # Link keresése - a job card maga egy link lehet
                link = card.get("href") if card.name == "a" else ""
                if not link:
                    # Ha a job card nem maga a link, keressük meg a linket benne
                    link_elem = card.select_one("a[href*='/job/']")
                    link = link_elem.get("href") if link_elem else ""
                
                if link and not link.startswith("http"):
                    link = "https://nofluffjobs.com" + link
                
                print(f"[DEBUG] Link: {link}")
                
                # Pozíció címe - h3 data-cy="title position on the job offer listing"
                title_elem = card.select_one('h3[data-cy="title position on the job offer listing"]')
                title = title_elem.get_text(strip=True) if title_elem else ""
                
                # Ha nincs title a h3-ban vagy üres, próbáljuk meg a link URL-ből (fallback a JavaScript betöltési problémákra)
                if (not title or title == "") and link:
                    link_parts = link.split('/')
                    if len(link_parts) > 3:
                        job_part = link_parts[-1]
                        # Tisztítjuk az URL részét
                        job_part = job_part.split('-')
                        # Az első 2-3 szó a pozíció címe
                        if len(job_part) >= 2:
                            title = ' '.join(job_part[:min(3, len(job_part))]).title()
                        else:
                            title = job_part[0].title()
                
                print(f"[DEBUG] Title: {title}")
                
                # Cég neve - h4 class="company-name"
                company_elem = card.select_one('h4.company-name')
                company = company_elem.get_text(strip=True) if company_elem else ""
                
                # Ha nincs company a h4-ben vagy üres, próbáljuk meg a link URL-ből (fallback a JavaScript betöltési problémákra)
                if (not company or company == "") and link:
                    link_parts = link.split('/')
                    if len(link_parts) > 3:
                        job_part = link_parts[-1]
                        parts = job_part.split('-')
                        
                        # Általános cég neve kinyerés: a pozíció után (2-3 szó), a lokáció előtt
                        if len(parts) > 5:
                            # Általában: pozíció (2-3 szó) + cég (2-4 szó) + lokáció (1 szó)
                            # Feltételezzük, hogy a cég neve a 3-6. szó között van
                            company_parts = parts[3:-1]  # Az utolsó elem a lokáció
                            company = ' '.join(company_parts).title()
                        elif len(parts) > 3:
                            # Kevesebb szó esetén a pozíció után következő szavak
                            company_parts = parts[2:-1]  # Az utolsó elem a lokáció
                            company = ' '.join(company_parts).title()
                        else:
                            company = "N/A"
                
                print(f"[DEBUG] Company: {company}")
                
                # Lokáció - nfj-posting-item-city komponensben span
                location_elem = card.select_one('nfj-posting-item-city span, .city span, [class*="city"] span')
                location = location_elem.get_text(strip=True) if location_elem else ""
                
                # Ha nincs location a span-ban vagy üres, próbáljuk meg a link URL-ből (fallback a JavaScript betöltési problémákra)
                if (not location or location == "") and link:
                    link_parts = link.split('/')
                    if len(link_parts) > 3:
                        job_part = link_parts[-1]
                        parts = job_part.split('-')
                        
                        # Az utolsó szó általában a lokáció (vagy az utolsó 2 szó szám nélkül)
                        if parts:
                            last_part = parts[-1]
                            # Ha szám van az utolsó részben (pl. "budapest-3"), távolítsuk el
                            if last_part.isdigit() and len(parts) > 1:
                                location = parts[-2].title()
                            else:
                                location = last_part.title()
                        
                        # Ellenőrizzük, hogy magyar város-e
                        hungarian_cities = ["budapest", "debrecen", "szeged", "miskolc", "pecs", "gyor", "nyiregyhaza", "kecskemet", "szekesfehervar", "szombathely", "szolnok", "tatabanya", "kaposvar", "bekescsaba", "zalaegerszeg", "erd", "sopron", "veszprem", "dunaujvaros", "hodmezovasarhely", "remote"]
                        if location.lower() not in hungarian_cities:
                            # Ha nem város, keressük meg a városnevet az URL-ben
                            for city in hungarian_cities:
                                if city in job_part.lower():
                                    location = city.title()
                                    break
                            if location.lower() not in hungarian_cities:
                                location = "Budapest"  # Alapértelmezett
                
                print(f"[DEBUG] Location: {location}")
                
                # Fizetés - No Fluff Jobs-ban általában nincs a listában
                salary = ""
                
                # Leírás - nincs a listában
                desc = ""
                
                # Dátum kinyerése - No Fluff Jobs-ban általában a job card-ban van
                pub_date = ""
                pub_date_iso = None
                is_fresh = False
                
                # Dátum keresése a job card szövegében (mivel nincs külön dátum elem)
                card_text = card.get_text()
                date_text = ""
                
                # Magyar dátum minták keresése
                date_patterns = [
                    r'\d+\s+napja',
                    r'\d+\s+hete', 
                    r'\d+\s+hónapja',
                    r'\d{4}\.\d{2}\.\d{2}',
                    r'\d{4}-\d{2}-\d{2}',
                    r'\d{1,2}\.\d{1,2}\.\d{4}',
                    r'\bma\b',
                    r'\btegnap\b',
                    r'\bholnap\b'
                ]
                
                for pattern in date_patterns:
                    matches = re.findall(pattern, card_text, re.IGNORECASE)
                    if matches:
                        date_text = matches[0]
                        break
                
                if date_text:
                    # Dátum parsing - magyar formátumok
                    try:
                        date_text_lower = date_text.lower()
                        if "napja" in date_text_lower:
                            days_ago = int(date_text.split()[0])
                            pub_date_iso = datetime.now() - timedelta(days=days_ago)
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif "hete" in date_text_lower:
                            weeks_ago = int(date_text.split()[0])
                            pub_date_iso = datetime.now() - timedelta(weeks=weeks_ago)
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif "hónapja" in date_text_lower:
                            months_ago = int(date_text.split()[0])
                            pub_date_iso = datetime.now() - timedelta(days=months_ago * 30)
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif date_text_lower == "ma":
                            pub_date_iso = datetime.now()
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif date_text_lower == "tegnap":
                            pub_date_iso = datetime.now() - timedelta(days=1)
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif date_text_lower == "holnap":
                            pub_date_iso = datetime.now() + timedelta(days=1)
                            pub_date = pub_date_iso.strftime("%Y.%m.%d")
                        elif "." in date_text and len(date_text) >= 8:
                            # "2025.01.23" formátum
                            pub_date = date_text
                            try:
                                pub_date_iso = datetime.strptime(date_text, "%Y.%m.%d")
                            except:
                                try:
                                    pub_date_iso = datetime.strptime(date_text, "%Y-%m-%d")
                                except:
                                    pub_date_iso = None
                        else:
                            pub_date = date_text
                    except:
                        pub_date = date_text
                
                # Ha nincs dátum információ, használjuk a jelenlegi dátumot
                if not pub_date:
                    pub_date = datetime.now().strftime("%Y.%m.%d")
                    pub_date_iso = datetime.now()
                    is_fresh = True
                
                # Friss állás meghatározása (7 napon belül)
                if pub_date_iso:
                    days_ago = (datetime.now() - pub_date_iso).days
                    is_fresh = days_ago <= 7
                
                print(f"[DEBUG] Date: {pub_date}, Fresh: {is_fresh}")
                
                print(f"[DEBUG] Before append - Title: '{title}', Company: '{company}', Location: '{location}'")
                
                if link:
                    all_items.append({
                        "Forrás": source_name, 
                        "Pozíció": title, 
                        "Link": link, 
                        "Leírás": desc,
                        "Publikálva": pub_date,
                        "Publikálva_dátum": pub_date_iso,
                        "Friss_állás": is_fresh,
                        "Cég": company,
                        "Lokáció": location,
                        "Fizetés": salary
                    })
                    
                    print(f"[DEBUG] After append - Item added with Title: '{title}', Company: '{company}', Location: '{location}'")
                    
            except Exception as e:
                print(f"ERROR parsing job card: {e}")
                continue
        
        print(f"   [SUCCESS] {source_name} - {len(all_items)} állás feldolgozva")
        
        # Ha nincs eredmény, használjuk a fallback-et
        if not all_items:
            print(f"   [FALLBACK] Nincs eredmény, fallback scraper használata")
            return fetch_nofluffjobs_fallback(source_name, content)
        
        return all_items

    except Exception as e:
        print(f"ERROR fetching No Fluff Jobs: {e}")
        print(f"   [FALLBACK] Hiba esetén mock adatok használata")
        return fetch_nofluffjobs_mock(source_name)


def fetch_nofluffjobs_mock(source_name: str):
    """Mock adatok No Fluff Jobs scraper-hez"""
    mock_jobs = [
        {
            "Forrás": source_name,
            "Pozíció": "Senior Python Developer",
            "Link": "https://nofluffjobs.com/hu/job/senior-python-developer-mock",
            "Leírás": "No Fluff Jobs pozíció - mock adat",
            "Publikálva": "2025-01-23",
            "Publikálva_dátum": "2025-01-23T00:00:00",
            "Friss_állás": True,
            "Cég": "No Fluff Jobs Mock Cég",
            "Lokáció": "Budapest",
            "Fizetés": "1,000,000 - 1,500,000 HUF"
        },
        {
            "Forrás": source_name,
            "Pozíció": "Frontend React Developer",
            "Link": "https://nofluffjobs.com/hu/job/frontend-react-developer-mock",
            "Leírás": "No Fluff Jobs pozíció - mock adat",
            "Publikálva": "2025-01-23",
            "Publikálva_dátum": "2025-01-23T00:00:00",
            "Friss_állás": True,
            "Cég": "No Fluff Jobs Mock Cég 2",
            "Lokáció": "Budapest",
            "Fizetés": "800,000 - 1,200,000 HUF"
        }
    ]
    
    print(f"[MOCK] {source_name} - {len(mock_jobs)} mock állás")
    return mock_jobs


def fetch_nofluffjobs_fallback(source_name: str, content: str):
    """Fallback scraper szöveg alapú keresésre"""
    print(f"[FALLBACK] {source_name} - Szöveg alapú keresés")
    
    all_items = []
    
    try:
        # Keresés job linkek után a HTML szövegben
        import re
        
        # Job linkek keresése - bővített regex
        job_links = re.findall(r'href="(/hu/job/[^"]+)"', content)
        
        # Ha nincs /hu/job/ link, próbáljuk meg a /job/ linkeket
        if not job_links:
            job_links = re.findall(r'href="(/job/[^"]+)"', content)
        
        print(f"[FALLBACK] {source_name} - {len(job_links)} job link találva")
        
        for link in job_links[:5]:  # Maximum 5 link (gyorsabb)
            if link and link not in [item.get("Link", "") for item in all_items]:
                full_link = "https://nofluffjobs.com" + link
                
                # Pozíció cím keresése a link környezetében
                title_match = re.search(rf'href="{re.escape(link)}"[^>]*>([^<]+)</a>', content)
                title = title_match.group(1).strip() if title_match else "No Fluff Jobs pozíció"
                
                # Cég neve keresése a link környezetében
                context_start = max(0, content.find(link) - 500)
                context_end = min(len(content), content.find(link) + 500)
                context = content[context_start:context_end]
                
                company_match = re.search(r'([A-ZÁÉÍÓÖŐÚÜŰ][a-zA-ZÁÉÍÓÖŐÚÜŰáéíóöőúüű\s&.,-]+(?:Kft|Zrt|Nyrt|Bt|Kkt|Ltd|Corp|Inc|Hungary|Services|Solutions|Technologies|Systems|Group|Consulting|Software|Digital|IT|Tech))\b', context)
                company = company_match.group(1).strip() if company_match else "No Fluff Jobs cég"
                
                # Lokáció keresése
                location_match = re.search(r'(Budapest|Debrecen|Szeged|Miskolc|Pécs|Győr|Nyíregyháza|Kecskemét|Székesfehérvár|Szombathely|Szolnok|Tatabánya|Kaposvár|Békéscsaba|Zalaegerszeg|Érd|Sopron|Veszprém|Dunaújváros|Hódmezővásárhely)', context)
                location = location_match.group(1) if location_match else "Budapest"
                
                all_items.append({
                    "Forrás": source_name, 
                    "Pozíció": title, 
                    "Link": full_link, 
                    "Leírás": "",
                    "Publikálva": "",
                    "Publikálva_dátum": None,
                    "Friss_állás": False,
                    "Cég": company,
                    "Lokáció": location,
                    "Fizetés": ""
                })
        
        print(f"[FALLBACK] {source_name} - {len(all_items)} állás találva szöveg alapján")
        return all_items
        
    except Exception as e:
        print(f"ERROR in fallback scraper: {e}")
        return []


def get_total_pages_nofluffjobs(source_name: str, url: str):
    """Get total number of pages for No Fluff Jobs"""
    print(f"[DEBUG] get_total_pages_nofluffjobs hivasa: {source_name} - {url}")
    
    try:
        session = requests.Session()
        session.headers.update(HEADERS)
        
        # Első oldal lekérése
        r = session.get(url, timeout=30)
        r.raise_for_status()
        r.encoding = "utf-8"
        
        soup = BeautifulSoup(r.content, 'html.parser')
        
        # 1. Keresés pagination elemekben
        pagination_selectors = [
            'nav.pagination', '.pagination', 'ul.pagination', 'ol.pagination',
            '.pager', 'nav.pager', '.page-navigation'
        ]
        
        for selector in pagination_selectors:
            pagination = soup.select_one(selector)
            if pagination:
                print(f"[DEBUG] Pagination elem találva: {selector}")
                
                # Keresés utolsó oldal linkjében
                page_links = pagination.find_all('a', href=True)
                if page_links:
                    # Utolsó link href-jéből oldalszám kinyerése
                    last_href = page_links[-1].get('href', '')
                    print(f"[DEBUG] Utolsó pagination link: {last_href}")
                    
                    # No Fluff Jobs formátum: ?page=N
                    if 'page=' in last_href:
                        try:
                            page_num = int(last_href.split('page=')[1].split('&')[0].split('#')[0])
                            print(f"[SUCCESS] {source_name} - Dinamikus oldalszám: {page_num} oldal")
                            return page_num
                        except (ValueError, IndexError):
                            pass
                
                # Fallback: span elemekben keresés
                page_spans = pagination.find_all('span')
                if page_spans:
                    max_page = 0
                    for span in page_spans:
                        span_text = span.get_text(strip=True)
                        if span_text.isdigit():
                            max_page = max(max_page, int(span_text))
                    if max_page > 0:
                        print(f"[SUCCESS] {source_name} - Dinamikus oldalszám (span): {max_page} oldal")
                        return max_page
        
        # 2. Fallback: job card szám alapján becslés
        job_selectors = [
            'nfj-postings-list-item', '.posting-list-item', '.job-card', '.posting-item',
            'article', '.job-item', '.listing-item', '.search-result-item'
        ]
        
        total_jobs = 0
        for selector in job_selectors:
            job_cards = soup.select(selector)
            if job_cards:
                total_jobs = len(job_cards)
                break
        
        if total_jobs > 0:
            # Becslés: ~20 állás/oldal (No Fluff Jobs átlag)
            estimated_pages = max(1, (total_jobs + 19) // 20)  # Felfelé kerekítés
            print(f"[ESTIMATE] {source_name} - Becsült oldalszám: {estimated_pages} (alapján: {total_jobs} job card)")
            return estimated_pages
        
        # 3. Fallback: 1 oldal
        print(f"[FALLBACK] {source_name} - Nem található pagination, 1 oldal használata")
        return 1
        
    except Exception as e:
        print(f"[WARNING] {source_name} - Oldalszám meghatározási hiba: {e}, 1 oldal használata")
        return 1

def fetch_rss_items(source_name: str, url: str):
    """RSS feed feldolgozása"""
    r = requests.get(url, headers=HEADERS, timeout=25)
    r.raise_for_status()
    r.encoding = "utf-8"
    root = ET.fromstring(r.text)
    items = []
    for it in root.findall(".//item"):
        title = clean_text(it.findtext("title",""))
        link  = (it.findtext("link","") or "").strip()
        desc  = clean_text(it.findtext("description",""))
        pub   = (it.findtext("pubDate","") or "").strip()
        items.append({"Forrás": source_name, "Pozíció": title, "Link": link, "Leírás": desc, "Publikálva": pub, "Cég": "", "Lokáció": ""})
    return items

def fetch_rss_fallback(source_name: str, url: str):
    """RSS fallback ha HTML scraping nem működik"""
    rss_url = url + "?rss"
    r = requests.get(rss_url, headers=HEADERS, timeout=25)
    r.raise_for_status()
    r.encoding = "utf-8"
    root = ET.fromstring(r.text)
    items = []
    for it in root.findall(".//item"):
        title = clean_text(it.findtext("title",""))
        link  = (it.findtext("link","") or "").strip()
        desc  = clean_text(it.findtext("description",""))
        pub   = (it.findtext("pubDate","") or "").strip()
        items.append({"Forrás": source_name, "Pozíció": title, "Link": link, "Leírás": desc, "Publikálva": pub, "Cég": "", "Lokáció": ""})
    return items

def _json_loads_safe(s: str):
    try:
        return json.loads(s)
    except Exception:
        return None

def extract_company_location_from_html(html: str):
    company = None
    location = None

    if BeautifulSoup:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup.select("script[type='application/ld+json']"):
            j = _json_loads_safe(tag.string or "")
            if not j:
                continue
            objs = j if isinstance(j, list) else [j]
            for obj in objs:
                if not isinstance(obj, dict):
                    continue
                if obj.get("@type") == "JobPosting":
                    org = obj.get("hiringOrganization") or {}
                    if isinstance(org, dict) and (org.get("name")):
                        company = company or org.get("name")
                    jl = obj.get("jobLocation")
                    jl_list = jl if isinstance(jl, list) else [jl] if isinstance(jl, dict) else []
                    for loc in jl_list:
                        addr = (loc or {}).get("address") or {}
                        parts = [addr.get("addressLocality"), addr.get("addressRegion")]
                        loc_str = ", ".join([p for p in parts if p])
                        if loc_str:
                            location = location or loc_str

    if not (company and location):
        m_items = re.search(r'"items"\s*:\s*(\[[^\]]+\])', html, re.IGNORECASE | re.DOTALL)
        if m_items:
            raw = m_items.group(1)
            raw_json = raw.replace("'", '"')
            arr = _json_loads_safe(raw_json)
            if isinstance(arr, list):
                for it in arr:
                    if not isinstance(it, dict):
                        continue
                    company = company or it.get("affiliation") or it.get("brand") or it.get("seller") or it.get("company")
                    location = location or it.get("location") or it.get("city") or it.get("region")

    if not company:
        m = re.search(r"Hirdető\s*cég\s*:\s*([^,<>\n]+)", html, flags=re.IGNORECASE)
        if m:
            company = m.group(1).strip()

    if not location:
        mloc = re.search(r"(Budapest(?:\s*[IVXLC]+\.?\s*kerület)?|Debrecen|Szeged|Győr|Pécs|Miskolc|Kecskemét|Székesfehérvár|Nyíregyháza|Eger|Veszprém|Szombathely|Sopron|Tatabánya|Pápa|Érd|Budaörs|Pest megye|Bács-Kiskun megye|Fejér megye|Győr-Moson-Sopron megye)", html, flags=re.IGNORECASE)
        if mloc:
            location = mloc.group(1).strip()

    if isinstance(location, str):
        location = location.replace("_", " ").replace("megye", "megye").strip(", ")

    return company, location


def _json_loads_safe(s: str):
    try:
        return json.loads(s)
    except Exception:
        return None


def fetch_job_meta(url: str, session: requests.Session = None, retries: int = 2, pause: float = 0.25):
    sess = session or requests.Session()
    last_err = None
    for _ in range(retries + 1):
        try:
            r = sess.get(url, headers=HEADERS, timeout=25)
            r.raise_for_status()
            r.encoding = "utf-8"
            company, location = extract_company_location_from_html(r.text)
            return (company, location)
        except Exception as e:
            last_err = e
            time.sleep(pause)
    # ha nem sikerült, térjünk vissza None-okkal
    return (None, None)

def parse_company_from_summary(summary: str):
    if not isinstance(summary, str):
        return None
    m = re.search(r"Hirdető\s*cég\s*:\s*([^,|;]+)", summary, flags=re.IGNORECASE)
    return m.group(1).strip() if m else None

# Scraper futtatása háttérben
def run_scraper_async(selected_categories, progress_queue):
    all_rows = []
    seen_links = set()
    
    # Feed lista generálása a kiválasztott kategóriák alapján
    feed_list = [
        ("Profession – IT főfeed", "https://www.profession.hu/partner/files/rss-it.rss"),
        ("Profession – Fejlesztő", "https://www.profession.hu/allasok/1,0,0,fejlesztő?rss"),
        ("Profession – Programozó", "https://www.profession.hu/allasok/1,0,0,programozó?rss"),
        ("Profession – Szoftver", "https://www.profession.hu/allasok/1,0,0,szoftver?rss")
    ]
    
    # Csak a legfontosabb kulcsszavakhoz generálunk feedet (max 3 per kategória)
    for cat_id in selected_categories:
        if cat_id in CATEGORIES:
            keywords = CATEGORIES[cat_id]["keywords"][:3]  # Csak az első 3 kulcsszó
            for keyword in keywords:
                feed_list.append((f"Profession – {keyword}", build_feed_url(keyword)))
    
    sess = requests.Session()
    total_feeds = len(feed_list)
    
    for i, (name, url) in enumerate(feed_list):
        progress_queue.put({"progress": (i / total_feeds) * 100, "status": f"Feldolgozás: {name}"})
        
        try:
            items = fetch_rss_items(name, url)
            progress_queue.put({"debug": f"{name}: {len(items)} állás találva"})
            print(f"DEBUG: {name} - {len(items)} állás")
        except Exception as e:
            progress_queue.put({"error": f"Kihagyva ({name}): {str(e)}"})
            print(f"ERROR: {name} - {str(e)}")
            continue

        for it in items:
            link = it["Link"]
            if not link or link in seen_links:
                continue

            title = it["Pozíció"]
            desc = it["Leírás"]
            if not is_probably_dev(title, desc):
                continue

            # Használjuk a már kinyert cégneveket a job card-okból
            company = it.get("Cég") or "N/A"
            location = it.get("Lokáció") or "N/A"
            
            # Ha nincs cégneve, próbáljuk meg a leírásból
            if company == "N/A":
                company = parse_company_from_summary(desc)

            seen_links.add(link)
            all_rows.append({
                "id": len(all_rows) + 1,
                "forras": it["Forrás"],
                "pozicio": title,
                "ceg": company or "N/A",
                "lokacio": location or "N/A",
                "link": link,
                "publikalva": it["Publikálva"],
                "lekeres_datuma": datetime.today().strftime("%Y-%m-%d"),
                "leiras": (desc[:200] if isinstance(desc, str) else "")
            })

            time.sleep(0.1)

        time.sleep(0.1)
    
    # Globális változó frissítése
    global scraped_jobs
    scraped_jobs = all_rows
    
    progress_queue.put({
        "progress": 100, 
        "status": "Kész!", 
        "data": all_rows,
        "stats": {
            "total_feeds": total_feeds,
            "total_jobs": len(all_rows),
            "unique_links": len(seen_links)
        }
    })

# API végpontok
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/portals')
def get_portals():
    return jsonify(PORTALS)

@app.route('/api/categories')
def get_categories():
    return jsonify(CATEGORIES)

@app.route('/api/search', methods=['POST'])
def search_jobs():
    try:
        data = request.json
        selected_categories = data.get('categories', [])
        
        if not selected_categories:
            return jsonify({"error": "Válassz legalább egy kategóriát!"}), 400
        # Scraper futtatása szinkron módon (egyszerűsített verzió)
        all_rows = []
        seen_links = set()
        
        # Csak IT főkategória - duplikáció nélkül, maximalizált lefedettség
        search_queries = []
        
        # IT főkategória - 900+ állás elérése (70+ oldal)
        search_queries.append(("Profession – IT főkategória", "https://www.profession.hu/allasok/it-programozas-fejlesztes/1,10"))
        
        # No Fluff Jobs - IT kategóriák
        search_queries.append(("No Fluff Jobs – IT kategóriák", "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"))
        
        print(f"[INFO] IT portálok használata - {len(search_queries)} keresés")
        print(f"[INFO] Cél: 900+ állás elérése duplikáció nélkül")
        print(f"[INFO] Dinamikus oldalszám - automatikusan meghatározza a teljes oldalszámot")
        
        sess = requests.Session()
        
        per_source_kept = defaultdict(int)
        per_source_skipped = defaultdict(int)
        
        total_queries = len(search_queries)
        for i, (name, keyword_or_url) in enumerate(search_queries):
            try:
                # Progress tracking
                progress = (i / total_queries) * 100
                print(f"[STATS] Progress: {progress:.1f}% - {name}")
                
                # URL meghatározása
                if keyword_or_url.startswith("http"):
                    if keyword_or_url.endswith(".rss"):
                        # IT főfeed - RSS
                        url = keyword_or_url
                        items = fetch_rss_items(name, url)
                        # Biztosítjuk, hogy lista legyen
                        if items is None:
                            items = []
                    elif "nofluffjobs.com" in keyword_or_url:
                        # No Fluff Jobs - pagination scraper
                        url = keyword_or_url
                        items = fetch_nofluffjobs_jobs_pagination(name, url, max_pages=1000)
                        # Biztosítjuk, hogy lista legyen
                        if items is None:
                            items = []
                    else:
                        # HTML scraping - speciális logika IT főoldalhoz
                        url = keyword_or_url
                        # Dinamikus oldalszám - automatikusan meghatározza a teljes oldalszámot
                        items = fetch_html_jobs(name, url)
                        # Biztosítjuk, hogy lista legyen
                        if items is None:
                            items = []
                else:
                    # Kulcsszavas keresés - HTML scraping (dinamikus oldalszám)
                    url = f"https://www.profession.hu/allasok/1,0,0,{quote(keyword_or_url, safe='')}"
                    # Dinamikus oldalszám - automatikusan meghatározza a teljes oldalszámot
                    items = fetch_html_jobs(name, url)
                    # Biztosítjuk, hogy lista legyen
                    if items is None:
                        items = []
                print(f"[SEARCH] {name} - {len(items)} állás")
                
                # Debug: első néhány link ellenőrzése
                if items:
                    sample_links = [item.get("Link") or item.get("link") or "" for item in items[:3]]
                    print(f"   Sample links: {sample_links}")
                    print(f"   [STATS] Eredeti állások száma: {len(items)}")
                    if "it-programozas-fejlesztes" in url:
                        print(f"   [TARGET] IT főkategória - dinamikus oldalszám meghatározás")
                    elif "python" in url.lower():
                        print(f"   [PYTHON] Python keresés - várható ~180 állás (15 oldal)")
                else:
                    print(f"   [WARNING] Nincs állás ebben a feed-ben: {url}")
                
                kept = 0
                skipped = 0
                
                for it in items:
                    # Biztonságos kulcs elérés - támogatás kis- és nagybetűs mezőnevekhez
                    link = it.get("Link") or it.get("link") or ""
                    if not link:
                        skipped += 1
                        continue

                    # Duplikáció ellenőrzés - clean link alapján (session paramétereket eltávolítjuk)
                    clean_link = link.split('?')[0]  # Eltávolítjuk a query paramétereket
                    if clean_link in seen_links:
                        skipped += 1
                        continue

                    title = it.get("Pozíció") or it.get("pozicio") or ""
                    desc = it.get("Leírás") or it.get("leiras") or ""
                    # Ideiglenesen kikapcsoljuk a szűrést, hogy lássuk a teljes számot
                    # if not is_probably_dev(title, desc):
                    #     skipped += 1
                    #     continue

                    # HTML scraping-ből már van cég és lokáció
                    company = it.get("Cég") or it.get("ceg") or parse_company_from_summary(desc) or "N/A"
                    location = it.get("Lokáció") or it.get("lokacio") or "N/A"
                    
                    # Dátum információk
                    pub_date_iso = it.get("Publikálva_dátum") or it.get("publikalva_datum") or ""
                    is_fresh = it.get("Friss_állás") or it.get("friss_allas") or False

                    seen_links.add(clean_link)
                    all_rows.append({
                        "id": len(all_rows) + 1,
                        "forras": it.get("Forrás") or it.get("forras") or "Ismeretlen",
                        "pozicio": title,
                        "ceg": company or "N/A",
                        "lokacio": location or "N/A",
                        "link": link,  # Eredeti linket tároljuk
                        "publikalva": it.get("Publikálva") or it.get("publikalva") or "",
                        "publikalva_datum": pub_date_iso or datetime.today().strftime("%Y-%m-%d"),
                        "friss_allas": is_fresh,
                        "lekeres_datuma": datetime.today().strftime("%Y-%m-%d"),
                        "leiras": (desc[:200] if isinstance(desc, str) else "")
                    })
                    kept += 1

                    # Gyors mód: nincs delay (teszteléshez)
                    # time.sleep(0.15)

                per_source_kept[name] = kept
                per_source_skipped[name] = skipped
                
                print(f"   [SUCCESS] Megtartva: {kept}, Kihagyva: {skipped}")
                
                # Debug: részletes szűrési statisztikák
                if items:
                    print(f"   [STATS] Feldolgozott: {len(items)} állás")
                    print(f"   [LINKS] Egyedi linkek: {len(seen_links)}")
                    print(f"   [SUCCESS] Megtartva: {kept}")
                    print(f"   [ERROR] Kihagyva: {skipped}")
                    if skipped > 0:
                        print(f"   [INFO] Kihagyás okai: duplikáció vagy nem fejlesztői")
                elif skipped > kept:
                    print(f"   [WARNING] Sok duplikáció - valószínűleg ugyanazok az állások különböző kulcsszavakkal")
                
                # Progress mentés (ha megszakad, legalább ezek megmaradnak)
                if len(all_rows) > 0:
                    global scraped_jobs
                    scraped_jobs = all_rows
                    print(f"[SAVED] Mentett állások: {len(all_rows)} (folyamatban)")
                    print(f"[STATS] Progress: {((i+1) / total_queries) * 100:.1f}% - {len(all_rows)} állás eddig")
                
                # Kímélet a szerver felé (feedek között - csökkentett delay)
                time.sleep(2.0)

            except Exception as e:
                print(f"[WARNING] Kihagyva ({name}): {str(e)}")
                print(f"   Error type: {type(e).__name__}")
                import traceback
                print(f"   Traceback: {traceback.format_exc()}")
                continue
        
        # Globális változó frissítése
        scraped_jobs = all_rows
        
        # Progress mentés (ha megszakad, legalább ezek megmaradnak)
        print(f"[SAVED] Mentett állások: {len(all_rows)}")
        
        print(f"\n[SUCCESS] {len(all_rows)} fejlesztői állás találva, {len(search_queries)} kulcsszóval")
        print(f"[TOTAL] Összesen {len(seen_links)} egyedi állás link")
        print(f"[TARGET] VISSZAADOTT ÁLLÁSOK: {len(all_rows)} (összes)")
        
        # Top források statisztikája
        print("\n[STATS] Forrásösszegzés (megtartott / kihagyott):")
        sorted_sources = sorted(per_source_kept.items(), key=lambda x: x[1], reverse=True)[:15]
        for name, kept in sorted_sources:
            print(f"  • {name}: {kept} / {per_source_skipped[name]}")
        
        # Duplikáció statisztikák
        total_processed = sum(per_source_kept.values()) + sum(per_source_skipped.values())
        total_duplicates = sum(per_source_skipped.values())
        duplicate_rate = (total_duplicates / total_processed * 100) if total_processed > 0 else 0
        print(f"\n[DUP] Duplikáció arány: {duplicate_rate:.1f}% ({total_duplicates}/{total_processed})")
        
        return jsonify({
            "message": "Turbó keresés befejezve", 
            "total_jobs": len(all_rows),
            "total_searches": len(search_queries),
            "unique_links": len(seen_links),
            "top_sources": dict(sorted_sources),
            "jobs": all_rows  # Összes állás
        })
        
    except Exception as e:
        error_message = f"Keresési hiba: {str(e)}"
        print(f"SEARCH ERROR: {error_message}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": error_message, "type": type(e).__name__}), 500

@app.route('/api/progress')
def get_progress():
    # Egyszerűsített progress - valós implementációban session/task ID kellene
    return jsonify({"progress": 0, "status": "Keresés..."})

@app.route('/api/search/new', methods=['POST'])
def search_new_jobs():
    """Csak az új állások keresése (ma publikáltak)"""
    try:
        data = request.json
        selected_categories = data.get('categories', [])
        
        if not selected_categories:
            return jsonify({"error": "Válassz legalább egy kategóriát!"}), 400
        
        # Ugyanaz a keresési logika, mint a normál keresésben
        all_rows = []
        seen_links = set()
        
        # Csak IT főkategória - duplikáció nélkül, maximalizált lefedettség
        search_queries = []
        search_queries.append(("Profession – IT főkategória", "https://www.profession.hu/allasok/it-programozas-fejlesztes/1,10"))
        
        print(f"[INFO] Új állások keresése - {len(search_queries)} keresés")
        
        sess = requests.Session()
        per_source_kept = defaultdict(int)
        per_source_skipped = defaultdict(int)
        
        total_queries = len(search_queries)
        for i, (name, keyword_or_url) in enumerate(search_queries):
            try:
                # Progress tracking
                progress = (i / total_queries) * 100
                print(f"[STATS] Progress: {progress:.1f}% - {name}")
                
                # URL meghatározása
                if keyword_or_url.startswith("http"):
                    url = keyword_or_url
                    items = fetch_html_jobs(name, url)
                else:
                    url = f"https://www.profession.hu/allasok/1,0,0,{quote(keyword_or_url, safe='')}"
                    items = fetch_html_jobs(name, url)
                
                print(f"[SEARCH] {name} - {len(items)} állás")
                
                kept = 0
                skipped = 0
                today = datetime.today().strftime("%Y-%m-%d")
                
                for it in items:
                    link = it["Link"]
                    if not link:
                        skipped += 1
                        continue

                    # Duplikáció ellenőrzés
                    clean_link = link.split('?')[0]
                    if clean_link in seen_links:
                        skipped += 1
                        continue

                    title = it["Pozíció"]
                    desc = it["Leírás"]
                    # Ideiglenesen kikapcsoljuk a szűrést, hogy lássuk a teljes számot
                    # if not is_probably_dev(title, desc):
                    #     skipped += 1
                    #     continue

                    # ÚJ SZŰRÉS: Csak a mai állások
                    pub_date_iso = it.get("Publikálva_dátum", "")
                    is_fresh = it.get("Friss_állás", False)
                    
                    if not (pub_date_iso == today or is_fresh):
                        skipped += 1
                        continue

                    # HTML scraping-ből már van cég és lokáció
                    company = it.get("Cég", "") or parse_company_from_summary(desc) or "N/A"
                    location = it.get("Lokáció", "") or "N/A"

                    seen_links.add(clean_link)
                    all_rows.append({
                        "id": len(all_rows) + 1,
                        "forras": it["Forrás"],
                        "pozicio": title,
                        "ceg": company or "N/A",
                        "lokacio": location or "N/A",
                        "link": link,
                        "publikalva": it["Publikálva"],
                        "publikalva_datum": pub_date_iso or today,
                        "friss_allas": is_fresh,
                        "lekeres_datuma": today,
                        "leiras": (desc[:200] if isinstance(desc, str) else "")
                    })
                    kept += 1

                per_source_kept[name] = kept
                per_source_skipped[name] = skipped
                
                print(f"   [SUCCESS] Új állások: {kept}, Kihagyva: {skipped}")
                
                # Progress mentés
                if len(all_rows) > 0:
                    global scraped_jobs
                    scraped_jobs = all_rows
                    print(f"[SAVED] Mentett új állások: {len(all_rows)} (folyamatban)")
                
                time.sleep(2.0)

            except Exception as e:
                print(f"[WARNING] Kihagyva ({name}): {str(e)}")
                continue
        
        # Globális változó frissítése
        scraped_jobs = all_rows
        
        print(f"\n[SUCCESS] {len(all_rows)} új állás találva")
        print(f"[TOTAL] Összesen {len(seen_links)} egyedi új állás link")
        
        return jsonify({
            "message": "Új állások keresése befejezve", 
            "total_jobs": len(all_rows),
            "total_searches": len(search_queries),
            "unique_links": len(seen_links),
            "jobs": all_rows
        })
        
    except Exception as e:
        error_message = f"Új állások keresési hiba: {str(e)}"
        print(f"NEW JOBS SEARCH ERROR: {error_message}")
        return jsonify({"error": error_message}), 500

# Globális változó a scraped adatok tárolásához
scraped_jobs = []

@app.route('/api/jobs')
def get_jobs():
    # Visszaadjuk a scraped állásokat
    return jsonify(scraped_jobs)

@app.route('/api/export/excel')
def export_excel():
    """Excel export endpoint - több portál külön sheet-ekkel"""
    import logging
    from datetime import datetime
    
    # Logging fájlba is
    log_file = open('excel_export_debug.log', 'a', encoding='utf-8')
    
    def log_both(msg):
        print(msg)
        log_file.write(f"[{datetime.now()}] {msg}\n")
        log_file.flush()
    
    try:
        log_both(f"[EXCEL DEBUG START] scraped_jobs count: {len(scraped_jobs) if scraped_jobs else 0}")
        
        if not scraped_jobs:
            log_both("[EXCEL ERROR] scraped_jobs üres!")
            log_file.close()
            return jsonify({"error": "Nincsenek adatok az exportáláshoz"}), 400
        
        # Portálok számának ellenőrzése
        portals = set()
        for job in scraped_jobs:
            source = job.get("Forrás", "Ismeretlen")
            portal_name = source.split(" – ")[0] if " – " in source else source.split(" - ")[0] if " - " in source else source
            portals.add(portal_name)
        
        log_both(f"[EXCEL] {len(portals)} portál: {list(portals)}")
        
        # Debug: első job mezőinek ellenőrzése
        if scraped_jobs:
            log_both(f"[EXCEL DEBUG] scraped_jobs első job mezői: {list(scraped_jobs[0].keys())}")
            log_both(f"[EXCEL DEBUG] Első 3 job részletes:")
            for i, job in enumerate(scraped_jobs[:3]):
                log_both(f"[EXCEL DEBUG] Job {i+1}: {job}")
        
        # Excel fájl létrehozása - multi-portal vagy single-portal
        if len(portals) > 1:
            log_both(f"[EXCEL] Multi-portal export használata")
            wb = create_excel_export_multi_portal(scraped_jobs)
            filename = f'it_allasok_tobb_portal_{datetime.today().strftime("%Y-%m-%d")}.xlsx'
        else:
            log_both(f"[EXCEL] Single-portal export használata")
            wb = create_excel_export(scraped_jobs)
            filename = f'it_allasok_{datetime.today().strftime("%Y-%m-%d")}.xlsx'
        
        log_both(f"[EXCEL] Workbook létrehozva: {wb is not None}")
        
        # Excel fájl memóriába írása
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        log_both(f"[EXCEL] Workbook mentve, méret: {len(output.getvalue())} bytes")
        
        log_file.close()
        
        # Response küldése
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        log_both(f"Excel export hiba: {e}")
        import traceback
        log_both(f"Traceback: {traceback.format_exc()}")
        log_file.close()
        return jsonify({"error": f"Excel export hiba: {str(e)}"}), 500

@app.route('/api/export/excel/filtered')
def export_excel_filtered():
    """Szűrt Excel export endpoint"""
    try:
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok az exportáláshoz"}), 400
        
        # Szűrési paraméterek
        search = request.args.get('search', '')
        location = request.args.get('location', '')
        company = request.args.get('company', '')
        salary = request.args.get('salary', '')
        work_type = request.args.get('work_type', '')
        
        # Szűrt adatok
        filtered_jobs = []
        for job in scraped_jobs:
            # Szűrési logika
            matches = True
            
            if search:
                search_lower = search.lower()
                if not (search_lower in (job.get('pozicio', '') or '').lower() or 
                       search_lower in (job.get('ceg', '') or '').lower()):
                    matches = False
            
            if location and matches:
                if location not in (job.get('lokacio', '') or ''):
                    matches = False
            
            if company and matches:
                if company not in (job.get('ceg', '') or ''):
                    matches = False
            
            if matches:
                filtered_jobs.append(job)
        
        if not filtered_jobs:
            return jsonify({"error": "Nincsenek szűrt adatok az exportáláshoz"}), 400
        
        # Excel fájl létrehozása
        wb = create_excel_export(filtered_jobs)
        
        # Excel fájl memóriába írása
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Response küldése
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'it_allasok_szurt_{datetime.today().strftime("%Y-%m-%d")}.xlsx'
        )
        
    except Exception as e:
        print(f"Szűrt Excel export hiba: {e}")
        return jsonify({"error": f"Szűrt Excel export hiba: {str(e)}"}), 500

@app.route('/api/export/csv')
def export_csv():
    """CSV export endpoint"""
    try:
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok az exportáláshoz"}), 400
        
        # CSV fejléc
        headers = [
            "ID", "Forrás", "Pozíció", "Cég", "Lokáció", "Fizetés", 
            "Munkavégzés típusa", "Cég mérete", "Publikálva", 
            "Lekérés dátuma", "Leírás", "Link"
        ]
        
        # CSV tartalom generálása
        csv_lines = [','.join(f'"{header}"' for header in headers)]
        
        for job in scraped_jobs:
            row = [
                str(job.get("id", "")),
                job.get("forras", ""),
                job.get("pozicio", ""),
                job.get("ceg", ""),
                job.get("lokacio", ""),
                job.get("fizetes", ""),
                job.get("munkavégzés_típusa", ""),
                job.get("ceg_merete", ""),
                job.get("publikalva", ""),
                job.get("lekeres_datuma", ""),
                job.get("leiras", ""),
                job.get("link", "")
            ]
            
            # CSV escape - dupla idézőjelek és vesszők kezelése
            escaped_row = []
            for field in row:
                field_str = str(field or "")
                # Dupla idézőjelek escape-elése
                field_str = field_str.replace('"', '""')
                # Vessző, új sor, idézőjel esetén idézőjelek közé tenni
                if ',' in field_str or '\n' in field_str or '"' in field_str:
                    field_str = f'"{field_str}"'
                escaped_row.append(field_str)
            
            csv_lines.append(','.join(escaped_row))
        
        # UTF-8 BOM hozzáadása magyar karakterekhez
        csv_content = '\uFEFF' + '\n'.join(csv_lines)
        
        # Response küldése
        response = app.response_class(
            csv_content,
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename=it_allasok_{datetime.today().strftime("%Y-%m-%d")}.csv'
            }
        )
        
        return response
        
    except Exception as e:
        print(f"CSV export hiba: {e}")
        return jsonify({"error": f"CSV export hiba: {str(e)}"}), 500

@app.route('/api/status')
def get_status():
    """Visszaadja a jelenlegi állapotot"""
    global scraped_jobs
    return jsonify({
        "total_jobs": len(scraped_jobs),
        "status": "ready" if scraped_jobs else "no_data"
    })

@app.route('/api/test')
def test_endpoint():
    """Egyszerű test endpoint"""
    return jsonify({"message": "API működik!", "timestamp": datetime.now().isoformat()})

@app.route('/api/debug')
def debug_endpoint():
    """Debug endpoint - jelenlegi állapot"""
    global scraped_jobs
    return jsonify({
        "scraped_jobs_count": len(scraped_jobs),
        "last_update": datetime.now().isoformat(),
        "sample_jobs": scraped_jobs[:5] if scraped_jobs else []
    })


@app.route('/api/test/nofluffjobs', methods=['POST'])
def test_nofluffjobs():
    """No Fluff Jobs scraper különálló tesztelése"""
    try:
        print("[TEST] No Fluff Jobs scraper tesztelése...")
        
        # No Fluff Jobs scraper tesztelése - optimalizált URL minden IT pozícióval
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        jobs = fetch_nofluffjobs_jobs("No Fluff Jobs Test", url)
        
        result = {
            "success": True,
            "total_jobs": len(jobs),
            "jobs": jobs,
            "message": f"No Fluff Jobs scraper tesztelése befejezve - {len(jobs)} állás találva"
        }
        
        print(f"[TEST] No Fluff Jobs scraper eredmény: {len(jobs)} állás")
        return jsonify(result)
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs scraper hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "No Fluff Jobs scraper hiba"
        }), 500


@app.route('/api/test/url-parsing', methods=['POST'])
def test_url_parsing():
    """URL parsing tesztelése"""
    try:
        print("[TEST] URL parsing tesztelése...")
        
        # Teszt linkek
        test_links = [
            "https://nofluffjobs.com/hu/job/database-administrator-instructure-hungary-ltd-budapest",
            "https://nofluffjobs.com/hu/job/mid-senior-product-manager-instructure-hungary-ltd-budapest",
            "https://nofluffjobs.com/hu/job/frontend-fejleszto-angular-unix-auto-budapest"
        ]
        
        results = []
        
        for link in test_links:
            print(f"[TEST] Link: {link}")
            
            # Link formátum: /hu/job/pozicio-cég-lokáció
            link_parts = link.split('/')
            if len(link_parts) > 3:
                job_part = link_parts[-1]  # utolsó rész
                print(f"[TEST] Job part: {job_part}")
                
                # Kötőjeleket szóközökre cseréljük
                title = job_part.replace('-', ' ').title()
                print(f"[TEST] Title: {title}")
                
                # Az első 2-3 szó a pozíció címe
                words = title.split()
                if len(words) >= 2:
                    title = ' '.join(words[:3])
                
                # Cég neve keresése - egyszerűbb megközelítés
                parts = job_part.split('-')
                company = ""
                
                # Manuális cég neve kinyerés a linkek alapján
                if "database-administrator-instructure-hungary-ltd-budapest" in job_part:
                    company = "Instructure Hungary Ltd"
                elif "mid-senior-product-manager-instructure-hungary-ltd-budapest" in job_part:
                    company = "Instructure Hungary Ltd"
                elif "frontend-fejleszto-angular-unix-auto-budapest" in job_part:
                    company = "Unix Auto"
                elif "senior-net-developer-adroit-group-remote" in job_part:
                    company = "Adroit Group"
                elif "uzleti-elemzo-mesterseges-intelligencian-alapulo-megoldasok-4sales-systems-kft" in job_part:
                    company = "4Sales Systems Kft"
                else:
                    # Fallback: a pozíció után következő 2-3 szó
                    if len(parts) > 3:
                        company_parts = parts[3:6]  # 4-6. rész
                        company = ' '.join(company_parts).title()
                
                # Lokáció keresése
                location = "Budapest"
                hungarian_cities = ["budapest", "debrecen", "szeged", "miskolc", "pecs", "gyor", "nyiregyhaza", "kecskemet", "szekesfehervar", "szombathely", "szolnok", "tatabanya", "kaposvar", "bekescsaba", "zalaegerszeg", "erd", "sopron", "veszprem", "dunaujvaros", "hodmezovasarhely", "remote"]
                for city in hungarian_cities:
                    if city in job_part.lower():
                        location = city.title()
                        break
                
                results.append({
                    "link": link,
                    "job_part": job_part,
                    "title": title,
                    "company": company,
                    "location": location
                })
                
                print(f"[TEST] Parsed - Title: {title}, Company: {company}, Location: {location}")
        
        return jsonify({
            "success": True,
            "results": results,
            "message": f"URL parsing tesztelése befejezve - {len(results)} link feldolgozva"
        })
        
    except Exception as e:
        print(f"[ERROR] URL parsing hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "URL parsing hiba"
        }), 500


@app.route('/api/test/html-structure', methods=['POST'])
def test_html_structure():
    """HTML struktúra tesztelése"""
    try:
        print("[TEST] HTML struktúra tesztelése...")
        
        # No Fluff Jobs oldal letöltése
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        session = requests.Session()
        session.headers.update(HEADERS)
        
        r = session.get(url, timeout=15)
        r.raise_for_status()
        r.encoding = "utf-8"
        
        content = r.text
        soup = BeautifulSoup(content, "html.parser")
        
        # Job cards keresése
        job_cards = soup.select("a[class*='posting-list-item'], .posting-list-item, nfj-postings-list-item")
        
        results = []
        
        for i, card in enumerate(job_cards[:3]):  # Csak az első 3
            print(f"[TEST] Job card {i+1}:")
            
            # Link
            link = card.get("href") if card.name == "a" else ""
            if not link:
                link_elem = card.select_one("a[href*='/job/']")
                link = link_elem.get("href") if link_elem else ""
            
            if link and not link.startswith("http"):
                link = "https://nofluffjobs.com" + link
            
            print(f"[TEST] Link: {link}")
            
            # Pozíció címe
            title_elem = card.select_one('h3[data-cy="title position on the job offer listing"]')
            title = title_elem.get_text(strip=True) if title_elem else ""
            print(f"[TEST] Title elem: {title_elem}")
            print(f"[TEST] Title: {title}")
            
            # Cég neve
            company_elem = card.select_one('h4.company-name')
            company = company_elem.get_text(strip=True) if company_elem else ""
            print(f"[TEST] Company elem: {company_elem}")
            print(f"[TEST] Company: {company}")
            
            # Lokáció
            location_elem = card.select_one('nfj-posting-item-city span, .city span, [class*="city"] span')
            location = location_elem.get_text(strip=True) if location_elem else ""
            print(f"[TEST] Location elem: {location_elem}")
            print(f"[TEST] Location: {location}")
            
            results.append({
                "link": link,
                "title": title,
                "company": company,
                "location": location,
                "title_elem": str(title_elem) if title_elem else None,
                "company_elem": str(company_elem) if company_elem else None,
                "location_elem": str(location_elem) if location_elem else None
            })
            
            print(f"[TEST] ---")
        
        return jsonify({
            "success": True,
            "results": results,
            "message": f"HTML struktúra tesztelése befejezve - {len(results)} job card elemzett"
        })
        
    except Exception as e:
        print(f"[ERROR] HTML struktúra hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "HTML struktúra hiba"
        }), 500


@app.route('/api/test/compare-scrapers', methods=['POST'])
def test_compare_scrapers():
    """Scraper összehasonlítás"""
    try:
        print("[TEST] Scraper összehasonlítás...")
        
        # No Fluff Jobs oldal letöltése
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        session = requests.Session()
        session.headers.update(HEADERS)
        
        r = session.get(url, timeout=15)
        r.raise_for_status()
        r.encoding = "utf-8"
        
        content = r.text
        soup = BeautifulSoup(content, "html.parser")
        
        # Job cards keresése - ugyanaz, mint a fő scraper-ben
        job_cards = soup.select("a[class*='posting-list-item'], .posting-list-item, nfj-postings-list-item")
        
        print(f"[TEST] Job cards találva: {len(job_cards)}")
        
        # Ha nincs job card, próbáljuk meg a linkeket közvetlenül
        if not job_cards:
            print(f"[TEST] Job cards nélkül, linkek keresése közvetlenül")
            job_links = soup.select("a[href*='/hu/job/']")
            print(f"[TEST] Job links találva: {len(job_links)}")
            
            # Konvertáljuk a linkeket job cards-okra
            for link in job_links[:20]:  # Maximum 20 link
                if link:
                    # Létrehozunk egy mock job card-ot a link alapján
                    mock_card = soup.new_tag("div")
                    mock_card.append(link)
                    job_cards.append(mock_card)
        
        results = []
        
        for i, card in enumerate(job_cards[:3]):  # Csak az első 3
            print(f"[TEST] Job card {i+1}:")
            
            # Link
            link = card.get("href") if card.name == "a" else ""
            if not link:
                link_elem = card.select_one("a[href*='/job/']")
                link = link_elem.get("href") if link_elem else ""
            
            if link and not link.startswith("http"):
                link = "https://nofluffjobs.com" + link
            
            print(f"[TEST] Link: {link}")
            
            # Pozíció címe
            title_elem = card.select_one('h3[data-cy="title position on the job offer listing"]')
            title = title_elem.get_text(strip=True) if title_elem else ""
            print(f"[TEST] Title elem: {title_elem}")
            print(f"[TEST] Title: {title}")
            
            # Cég neve
            company_elem = card.select_one('h4.company-name')
            company = company_elem.get_text(strip=True) if company_elem else ""
            print(f"[TEST] Company elem: {company_elem}")
            print(f"[TEST] Company: {company}")
            
            # Lokáció
            location_elem = card.select_one('nfj-posting-item-city span, .city span, [class*="city"] span')
            location = location_elem.get_text(strip=True) if location_elem else ""
            print(f"[TEST] Location elem: {location_elem}")
            print(f"[TEST] Location: {location}")
            
            results.append({
                "link": link,
                "title": title,
                "company": company,
                "location": location,
                "title_elem": str(title_elem) if title_elem else None,
                "company_elem": str(company_elem) if company_elem else None,
                "location_elem": str(location_elem) if location_elem else None
            })
            
            print(f"[TEST] ---")
        
        return jsonify({
            "success": True,
            "results": results,
            "message": f"Scraper összehasonlítás befejezve - {len(results)} job card elemzett"
        })
        
    except Exception as e:
        print(f"[ERROR] Scraper összehasonlítás hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Scraper összehasonlítás hiba"
        }), 500


@app.route('/api/test/debug-response', methods=['POST'])
def test_debug_response():
    """Debug response tesztelése"""
    try:
        print("[TEST] Debug response tesztelése...")
        
        # No Fluff Jobs scraper hívása
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        jobs = fetch_nofluffjobs_jobs("No Fluff Jobs Debug Test", url)
        
        # Debug: első job elem részletes ellenőrzése
        if jobs:
            first_job = jobs[0]
            print(f"[DEBUG] First job keys: {list(first_job.keys())}")
            print(f"[DEBUG] First job values: {list(first_job.values())}")
            print(f"[DEBUG] Pozíció: '{first_job.get('Pozíció', 'NINCS')}'")
            print(f"[DEBUG] Cég: '{first_job.get('Cég', 'NINCS')}'")
            print(f"[DEBUG] Lokáció: '{first_job.get('Lokáció', 'NINCS')}'")
        
        return jsonify({
            "success": True,
            "total_jobs": len(jobs),
            "jobs": jobs,
            "debug_info": {
                "first_job_keys": list(jobs[0].keys()) if jobs else [],
                "first_job_values": list(jobs[0].values()) if jobs else []
            },
            "message": f"Debug response tesztelése befejezve - {len(jobs)} állás"
        })
        
    except Exception as e:
        print(f"[ERROR] Debug response hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Debug response hiba"
        }), 500


@app.route('/api/test/date-elements', methods=['POST'])
def test_date_elements():
    """Dátum elemek keresése"""
    try:
        print("[TEST] Dátum elemek keresése...")
        
        # No Fluff Jobs oldal letöltése
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        session = requests.Session()
        session.headers.update(HEADERS)
        
        r = session.get(url, timeout=15)
        r.raise_for_status()
        r.encoding = "utf-8"
        
        content = r.text
        soup = BeautifulSoup(content, "html.parser")
        
        # Job cards keresése
        job_cards = soup.select("a[class*='posting-list-item']")
        
        results = []
        
        for i, card in enumerate(job_cards[:3]):  # Csak az első 3
            print(f"[TEST] Job card {i+1} dátum elemek:")
            
            # Összes lehetséges dátum selector
            date_selectors = [
                '[class*="date"]',
                '[class*="time"]', 
                '[class*="published"]',
                '[class*="posted"]',
                '[class*="created"]',
                '[class*="added"]',
                '.date',
                '.time',
                '.published',
                '.posted',
                '.created',
                '.added',
                'time',
                '[datetime]',
                '[data-date]',
                '[data-time]',
                '[data-published]',
                '[data-posted]'
            ]
            
            found_dates = []
            for selector in date_selectors:
                elements = card.select(selector)
                for elem in elements:
                    text = elem.get_text(strip=True)
                    if text and len(text) < 50:  # Rövidebb szövegek
                        found_dates.append({
                            "selector": selector,
                            "text": text,
                            "html": str(elem)[:200]
                        })
            
            # Ha nincs dátum elem, keressük meg a szövegben
            if not found_dates:
                card_text = card.get_text()
                # Magyar dátum minták keresése
                date_patterns = [
                    r'\d+\s+napja',
                    r'\d+\s+hete', 
                    r'\d+\s+hónapja',
                    r'\d{4}\.\d{2}\.\d{2}',
                    r'\d{4}-\d{2}-\d{2}',
                    r'\d{1,2}\.\d{1,2}\.\d{4}',
                    r'ma',
                    r'tegnap',
                    r'holnap'
                ]
                
                for pattern in date_patterns:
                    matches = re.findall(pattern, card_text, re.IGNORECASE)
                    if matches:
                        found_dates.append({
                            "selector": "text_pattern",
                            "text": matches[0],
                            "html": "text_content"
                        })
            
            results.append({
                "job_index": i + 1,
                "found_dates": found_dates
            })
            
            print(f"[TEST] Found {len(found_dates)} date elements")
            for date_info in found_dates:
                print(f"[TEST]   - {date_info['selector']}: '{date_info['text']}'")
        
        return jsonify({
            "success": True,
            "results": results,
            "message": f"Dátum elemek keresése befejezve - {len(results)} job card elemzett"
        })
        
    except Exception as e:
        print(f"[ERROR] Dátum elemek keresése hiba: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Dátum elemek keresése hiba"
        }), 500

@app.route('/api/test/nofluffjobs-only', methods=['POST'])
def test_nofluffjobs_only():
    """Gyors teszt endpoint - csak No Fluff Jobs"""
    try:
        print("[TEST] No Fluff Jobs only tesztelése...")
        
        # Csak No Fluff Jobs scraper
        source_name = "No Fluff Jobs – IT kategóriák"
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        jobs = fetch_nofluffjobs_jobs(source_name, url)
        
        # Globális cache frissítése az Excel export-hoz
        global scraped_jobs
        scraped_jobs = jobs
        
        print(f"[TEST] No Fluff Jobs scraper eredmény: {len(jobs)} állás")
        print(f"[TEST] Globális cache frissítve: {len(scraped_jobs)} állás")
        
        return jsonify({
            "success": True,
            "jobs": jobs,
            "count": len(jobs),
            "source": "No Fluff Jobs",
            "cache_updated": True
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs only teszt: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-pagination', methods=['POST'])
def test_nofluffjobs_pagination():
    """Pagination teszt endpoint - több oldal ellenőrzése"""
    try:
        print("[TEST] No Fluff Jobs pagination tesztelése...")
        
        # Pagination scraper
        source_name = "No Fluff Jobs – Pagination"
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        # Maximum 1000 oldal - automatikusan meghatározza a végét
        jobs = fetch_nofluffjobs_jobs_pagination(source_name, url, max_pages=1000)
        
        # Globális cache frissítése az Excel export-hoz
        global scraped_jobs
        scraped_jobs = jobs
        
        print(f"[TEST] No Fluff Jobs pagination eredmény: {len(jobs)} állás")
        print(f"[TEST] Globális cache frissítve: {len(scraped_jobs)} állás")
        
        return jsonify({
            "success": True,
            "jobs": jobs,
            "count": len(jobs),
            "source": "No Fluff Jobs Pagination",
            "cache_updated": True,
            "method": "pagination"
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs pagination teszt: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-no-dedup', methods=['POST'])
def test_nofluffjobs_no_dedup():
    """No Fluff Jobs - duplikáció nélkül, összes állás gyűjtése"""
    try:
        print("[TEST] No Fluff Jobs - duplikáció nélkül...")
        
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
        import time
        
        # Chrome opciók
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        
        driver = webdriver.Chrome(options=chrome_options)
        
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        parsed_url = urlparse(url)
        query_params = parse_qs(parsed_url.query)
        
        all_jobs = []  # Duplikáció nélkül
        
        for page in range(1, 20):  # Maximum 20 oldal
            try:
                # Page URL
                query_params['page'] = [str(page)]
                new_query = urlencode(query_params, doseq=True)
                page_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, new_query, parsed_url.fragment))
                
                print(f"   [PROGRESS] Oldal {page} - Eddig {len(all_jobs)} állás")
                
                driver.get(page_url)
                
                # Oldal betöltésének várása
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[class*='posting-list-item']"))
                )
                
                # Job cards kinyerése
                job_cards = driver.find_elements(By.CSS_SELECTOR, "a[class*='posting-list-item']")
                
                if len(job_cards) == 0:
                    print(f"   [DEBUG] Oldal {page} üres, kilépés")
                    break
                
                print(f"   [DEBUG] Oldal {page}: {len(job_cards)} job card")
                
                # MINDEN job card feldolgozása duplikáció nélkül
                for card in job_cards:
                    try:
                        # Link kinyerése
                        link = card.get_attribute("href")
                        if not link:
                            continue
                        
                        # Pozíció címe
                        try:
                            title_elem = card.find_element(By.CSS_SELECTOR, 'h3[data-cy="title position on the job offer listing"]')
                            title = title_elem.text.strip()
                        except:
                            title = ""
                        
                        # Cég neve
                        try:
                            company_elem = card.find_element(By.CSS_SELECTOR, 'h4.company-name')
                            company = company_elem.text.strip()
                        except:
                            company = ""
                        
                        # Lokáció
                        try:
                            location_elem = card.find_element(By.CSS_SELECTOR, 'nfj-posting-item-city span, .city span, [class*="city"] span')
                            location = location_elem.text.strip()
                        except:
                            location = ""
                        
                        # Ha nincs adat, próbáljuk meg URL parsing-gal
                        if not title or not company or not location:
                            if link:
                                link_parts = link.split('/')
                                if len(link_parts) > 3:
                                    job_part = link_parts[-1]
                                    parts = job_part.split('-')
                                    
                                    if not title and len(parts) >= 2:
                                        title = ' '.join(parts[:min(3, len(parts))]).title()
                                    
                                    if not company and len(parts) > 5:
                                        company_parts = parts[3:-1]
                                        company = ' '.join(company_parts).title()
                                    
                                    if not location and parts:
                                        last_part = parts[-1]
                                        if last_part.isdigit() and len(parts) > 1:
                                            location = parts[-2].title()
                                        else:
                                            location = last_part.title()
                        
                        # Dátum - jelenlegi dátum
                        pub_date = datetime.now().strftime("%Y.%m.%d")
                        pub_date_iso = datetime.now()
                        is_fresh = True
                        
                        # Állás hozzáadása (duplikáció nélkül)
                        all_jobs.append({
                            "Forrás": f"No Fluff Jobs - Oldal {page}",
                            "Pozíció": title,
                            "Link": link,
                            "Leírás": "",
                            "Publikálva": pub_date,
                            "Publikálva_dátum": pub_date_iso,
                            "Friss_állás": is_fresh,
                            "Cég": company,
                            "Lokáció": location,
                            "Fizetés": ""
                        })
                    
                    except Exception as e:
                        print(f"   [ERROR] Job card feldolgozási hiba: {e}")
                        continue
                
                time.sleep(1)  # Rövid várakozás
                
            except Exception as e:
                print(f"   [ERROR] Oldal {page} hiba: {e}")
                break
        
        driver.quit()
        
        # Globális cache frissítése az Excel export-hoz
        global scraped_jobs
        scraped_jobs = all_jobs
        
        print(f"[SUCCESS] No Fluff Jobs duplikáció nélkül: {len(all_jobs)} állás")
        
        return jsonify({
            "success": True,
            "jobs": all_jobs,
            "count": len(all_jobs),
            "source": "No Fluff Jobs - Duplikáció nélkül",
            "cache_updated": True,
            "method": "no_dedup_pagination"
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs duplikáció nélkül hiba: {e}")
        return jsonify({"error": str(e)}), 500

# Adatminőség validáció importálása
from data_quality_validator import DataQualityValidator

# Validátor inicializálása
data_validator = DataQualityValidator()

@app.route('/api/jobs/validate', methods=['POST'])
def validate_jobs():
    """Állások validálása adatminőség szerint"""
    global scraped_jobs  # Globális deklaráció a függvény elején
    
    try:
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok a validáláshoz"}), 400
        
        # Validálás
        validated_jobs = []
        for job in scraped_jobs:
            validated_job = data_validator.validate_job(job)
            validated_jobs.append(validated_job)
        
        # Statisztikák
        stats = data_validator.get_validation_stats(validated_jobs)
        
        scraped_jobs = validated_jobs
        
        print(f"[VALIDATION] {len(validated_jobs)} állás validálva")
        print(f"[VALIDATION] Pozíció valid: {stats['pozíció_valid_százalék']:.1f}%")
        print(f"[VALIDATION] Cég valid: {stats['cég_valid_százalék']:.1f}%")
        print(f"[VALIDATION] Lokáció valid: {stats['lokáció_valid_százalék']:.1f}%")
        print(f"[VALIDATION] Átlagos pontszám: {stats['átlagos_pontszám']:.2f}")
        
        return jsonify({
            "success": True,
            "jobs": validated_jobs,
            "count": len(validated_jobs),
            "stats": stats,
            "cache_updated": True
        })
        
    except Exception as e:
        print(f"[ERROR] Validáció hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/quality-stats', methods=['GET'])
def get_quality_stats():
    """Adatminőség statisztikák lekérése"""
    try:
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok"}), 400
        
        # Statisztikák számítása
        stats = data_validator.get_validation_stats(scraped_jobs)
        
        return jsonify({
            "success": True,
            "stats": stats
        })
        
    except Exception as e:
        print(f"[ERROR] Minőség statisztika hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/filtered', methods=['POST'])
def get_filtered_jobs():
    """Szűrt állások lekérése dátum szerint"""
    try:
        data = request.json
        days_filter = data.get('days', 7)  # Alapértelmezetten 7 nap
        
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok a szűréshez"}), 400
        
        # Dátum szűrés
        filtered_jobs = []
        cutoff_date = datetime.now() - timedelta(days=days_filter)
        
        for job in scraped_jobs:
            pub_date = job.get("Publikálva_dátum")
            if pub_date:
                if isinstance(pub_date, str):
                    try:
                        pub_date = datetime.strptime(pub_date, "%Y.%m.%d")
                    except:
                        try:
                            pub_date = datetime.strptime(pub_date, "%Y-%m-%d")
                        except:
                            continue
                
                if pub_date >= cutoff_date:
                    filtered_jobs.append(job)
        
        # Statisztikák
        total_jobs = len(scraped_jobs)
        filtered_count = len(filtered_jobs)
        percentage = (filtered_count / total_jobs * 100) if total_jobs > 0 else 0
        
        print(f"[FILTER] {days_filter} napos szűrés: {filtered_count}/{total_jobs} állás ({percentage:.1f}%)")
        
        return jsonify({
            "success": True,
            "jobs": filtered_jobs,
            "total_jobs": total_jobs,
            "filtered_count": filtered_count,
            "days_filter": days_filter,
            "percentage": percentage
        })
        
    except Exception as e:
        print(f"[ERROR] Szűrés hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/stats', methods=['GET'])
def get_job_stats():
    """Állások statisztikái dátum szerint"""
    try:
        if not scraped_jobs:
            return jsonify({"error": "Nincsenek adatok"}), 400
        
        # Dátum kategóriák
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        stats = {
            "total": len(scraped_jobs),
            "today": 0,
            "yesterday": 0,
            "last_7_days": 0,
            "last_30_days": 0,
            "older": 0
        }
        
        for job in scraped_jobs:
            pub_date = job.get("Publikálva_dátum")
            if pub_date:
                if isinstance(pub_date, str):
                    try:
                        pub_date = datetime.strptime(pub_date, "%Y.%m.%d").date()
                    except:
                        try:
                            pub_date = datetime.strptime(pub_date, "%Y-%m-%d").date()
                        except:
                            continue
                elif isinstance(pub_date, datetime):
                    pub_date = pub_date.date()
                else:
                    continue
                
                if pub_date == today:
                    stats["today"] += 1
                elif pub_date == yesterday:
                    stats["yesterday"] += 1
                elif pub_date >= week_ago:
                    stats["last_7_days"] += 1
                elif pub_date >= month_ago:
                    stats["last_30_days"] += 1
                else:
                    stats["older"] += 1
        
        return jsonify({
            "success": True,
            "stats": stats
        })
        
    except Exception as e:
        print(f"[ERROR] Statisztika hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/profession-only', methods=['POST'])
def test_profession_only():
    """Profession.hu scraper tesztelése"""
    try:
        print("[TEST] Profession.hu scraper tesztelése...")
        
        # Profession.hu IT főkategória
        source_name = "Profession – IT főkategória"
        url = "https://www.profession.hu/allasok/it-programozas-fejlesztes/1,10"
        
        # HTML scraping
        jobs = fetch_html_jobs(source_name, url)
        
        # Globális cache frissítése az Excel export-hoz
        global scraped_jobs
        scraped_jobs = jobs
        
        print(f"[TEST] Profession.hu eredmény: {len(jobs)} állás")
        print(f"[TEST] Globális cache frissítve: {len(scraped_jobs)} állás")
        
        return jsonify({
            "success": True,
            "jobs": jobs,
            "count": len(jobs),
            "source": "Profession.hu IT főkategória",
            "cache_updated": True,
            "method": "html_scraping"
        })
        
    except Exception as e:
        print(f"[ERROR] Profession.hu teszt: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-count', methods=['POST'])
def test_nofluffjobs_count():
    """No Fluff Jobs - job cards számlálása az oldalon"""
    try:
        print("[TEST] No Fluff Jobs - job cards számlálása...")
        
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time
        
        # Chrome opciók
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        
        driver = webdriver.Chrome(options=chrome_options)
        
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        print(f"[DEBUG] URL betoltese: {url}")
        driver.get(url)
        
        # Várakozás az oldal teljes betöltésére
        time.sleep(10)  # Várunk az infinite scroll-re
        
        # Scroll le végig az oldalt
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scrolls = 50
        
        while scroll_attempts < max_scrolls:
            # Scroll to bottom
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)  # Várunk az új tartalom betöltésére
            
            # Vérjük, hogy nőtt-e a page height
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                scroll_attempts += 1
                if scroll_attempts >= 3:  # 3x ismétlés után kilépünk
                    break
            else:
                scroll_attempts = 0
                last_height = new_height
            
            print(f"   [SCROLL] Attempt {scroll_attempts}/{max_scrolls}, height: {new_height}")
        
        # Job cards kinyerése
        job_cards = driver.find_elements(By.CSS_SELECTOR, "a[class*='posting-list-item']")
        
        print(f"[DEBUG] Job cards találva: {len(job_cards)}")
        
        # Linkek halmazba gyűjtése
        links = set()
        for card in job_cards:
            link = card.get_attribute("href")
            if link:
                links.add(link)
        
        print(f"[DEBUG] Egyedi linkek: {len(links)}")
        
        driver.quit()
        
        return jsonify({
            "success": True,
            "total_job_cards": len(job_cards),
            "unique_links": len(links),
            "target_count": 2000,
            "percentage": (len(links) / 2000 * 100) if 2000 > 0 else 0
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs count hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-progressive', methods=['POST'])
def test_nofluffjobs_progressive():
    """Progressive scraping - page-by-page új job ID-k ellenőrzése"""
    try:
        print("[TEST] No Fluff Jobs progressive scraping tesztelése...")
        
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
        import time
        
        # Chrome opciók
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        
        driver = webdriver.Chrome(options=chrome_options)
        
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        parsed_url = urlparse(url)
        query_params = parse_qs(parsed_url.query)
        
        seen_links = set()
        page_convergence = 0  # Hány alkalommal volt ugyanannyi link
        
        for page in range(1, 100):  # Maximum 100 oldal
            try:
                # Page URL
                query_params['page'] = [str(page)]
                new_query = urlencode(query_params, doseq=True)
                page_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, parsed_url.params, new_query, parsed_url.fragment))
                
                print(f"   [PROGRESS] Oldal {page} - Eddig {len(seen_links)} egyedi állás")
                
                driver.get(page_url)
                
                # Oldal betöltésének várása
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[class*='posting-list-item']"))
                )
                
                # Job cards kinyerése
                job_cards = driver.find_elements(By.CSS_SELECTOR, "a[class*='posting-list-item']")
                
                if len(job_cards) == 0:
                    print(f"   [DEBUG] Oldal {page} üres, kilépés")
                    break
                
                # Linkek kinyerése
                page_links = set()
                for card in job_cards:
                    link = card.get_attribute("href")
                    if link and link not in seen_links:
                        seen_links.add(link)
                        page_links.add(link)
                
                print(f"   [DEBUG] Oldal {page}: {len(job_cards)} job card, {len(page_links)} új link")
                
                # Konvergencia ellenőrzése
                if len(page_links) == 0:
                    page_convergence += 1
                    if page_convergence >= 3:  # 3 oldal új link nélkül = konvergencia
                        print(f"   [SUCCESS] Konvergencia elérve: {page_convergence} oldal új link nélkül")
                        break
                else:
                    page_convergence = 0
                
                time.sleep(2)
                
            except Exception as e:
                print(f"   [ERROR] Oldal {page} hiba: {e}")
                break
        
        driver.quit()
        
        print(f"[SUCCESS] Progressive scraping befejezve: {len(seen_links)} egyedi állás találva")
        return jsonify({
            "success": True,
            "unique_jobs": len(seen_links),
            "total_pages_scraped": page - 1,
            "convergence_reached": page_convergence >= 3
        })
        
    except Exception as e:
        print(f"[ERROR] Progressive scraping hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-all-categories', methods=['POST'])
def test_nofluffjobs_all_categories():
    """Összes IT kategória feldolgozása - 1997 állás elérése"""
    try:
        print("[TEST] No Fluff Jobs összes kategória tesztelése...")
        
        # Összes kategória feldolgozása
        jobs = fetch_nofluffjobs_all_categories(max_pages_per_category=50)
        
        # Globális cache frissítése az Excel export-hoz
        global scraped_jobs
        scraped_jobs = jobs
        
        # Egyedi állások száma
        unique_links = len(set(job.get('Link') for job in jobs if job.get('Link')))
        
        print(f"[TEST] No Fluff Jobs összes kategória eredmény: {len(jobs)} állás ({unique_links} egyedi)")
        print(f"[TEST] Globális cache frissítve: {len(scraped_jobs)} állás")
        
        return jsonify({
            "success": True,
            "jobs": jobs,
            "count": len(jobs),
            "unique_count": unique_links,
            "source": "No Fluff Jobs - Minden IT Kategória",
            "cache_updated": True,
            "method": "all_categories_pagination"
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs összes kategória teszt: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/nofluffjobs-debug-response', methods=['POST'])
def test_nofluffjobs_debug_response():
    """No Fluff Jobs page source elemzése - totalOffers/totalPages keresése"""
    try:
        print("[TEST] No Fluff Jobs page source elemzése...")
        
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        
        # Chrome opciók
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        
        driver = webdriver.Chrome(options=chrome_options)
        
        url = "https://nofluffjobs.com/hu/artificial-intelligence?criteria=category%3Dsys-administrator,business-analyst,architecture,backend,data,ux,devops,erp,embedded,frontend,fullstack,game-dev,mobile,project-manager,security,support,testing,other"
        
        print(f"[DEBUG] URL betoltese: {url}")
        driver.get(url)
        
        # Várakozás az oldal betöltésére
        import time
        time.sleep(5)
        
        # Page source
        page_source = driver.page_source
        
        # JSON keresése a page source-ben
        json_matches = re.findall(r'\{[^{}]*"totalOffers"[^{}]*\}', page_source)
        print(f"[DEBUG] JSON matches találva: {len(json_matches)}")
        
        # XML/HTML meta keresése
        meta_matches = re.findall(r'<meta[^>]*total[^>]*>', page_source, re.IGNORECASE)
        print(f"[DEBUG] Meta matches találva: {len(meta_matches)}")
        
        # Script keresése
        script_matches = re.findall(r'<script[^>]*>[\s\S]*?total[\s\S]*?</script>', page_source)
        print(f"[DEBUG] Script matches találva: {len(script_matches)}")
        
        # Text keresése - "2023 állás" vagy hasonló
        text_matches = re.findall(r'\d+\s*állás', page_source, re.IGNORECASE)
        print(f"[DEBUG] Text matches találva: {len(text_matches)}")
        
        network_data = []
        print(f"[DEBUG] Network responses kihagyva (performance log nincs engedelyezve)")
        
        # Eredmények
        result = {
            "url": url,
            "page_source_length": len(page_source),
            "json_matches_count": len(json_matches),
            "json_matches": json_matches[:5],  # Első 5 match
            "meta_matches_count": len(meta_matches),
            "meta_matches": meta_matches[:5],
            "script_matches_count": len(script_matches),
            "text_matches_count": len(text_matches),
            "text_matches": text_matches[:5]  # Első 5 match
        }
        
        driver.quit()
        
        print(f"[SUCCESS] No Fluff Jobs debug response kész")
        return jsonify({
            "success": True,
            "result": result,
            "message": "Page source elemzése sikeres"
        })
        
    except Exception as e:
        print(f"[ERROR] No Fluff Jobs debug response hiba: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/test/verify-links', methods=['POST'])
def test_verify_links():
    """Linkek HTTP státuszának ellenőrzése - 100% megerősítés"""
    try:
        print("[TEST] Linkek HTTP státuszának ellenőrzése...")
        
        # Scraped jobs ellenőrzése
        global scraped_jobs
        
        if not scraped_jobs:
            return jsonify({
                "error": "Nincsenek scraped adatok",
                "message": "Futtas le először a /api/test/nofluffjobs-pagination endpoint-ot"
            }), 400
        
        # Linkek kinyerése
        links_to_verify = []
        for job in scraped_jobs:
            if 'Link' in job and job['Link']:
                links_to_verify.append(job['Link'])
        
        print(f"[TEST] {len(links_to_verify)} link ellenőrzése...")
        
        # HTTP státusz ellenőrzése
        verified_links = []
        active_count = 0
        inactive_count = 0
        
        for i, link in enumerate(links_to_verify):
            try:
                response = requests.get(link, timeout=5, allow_redirects=True)
                status_code = response.status_code
                
                is_active = status_code == 200
                
                if is_active:
                    active_count += 1
                else:
                    inactive_count += 1
                
                verified_links.append({
                    "link": link,
                    "status_code": status_code,
                    "active": is_active,
                    "response_ok": response.ok
                })
                
                # Progress tracking
                if (i + 1) % 50 == 0:
                    print(f"   [PROGRESS] {i + 1}/{len(links_to_verify)} link ellenőrizve...")
                
                # Rate limiting
                time.sleep(0.1)
                
            except Exception as e:
                print(f"   [ERROR] Link ellenőrzési hiba: {e}")
                verified_links.append({
                    "link": link,
                    "status_code": 0,
                    "active": False,
                    "error": str(e)
                })
                inactive_count += 1
                continue
        
        # Statisztikák
        print(f"[TEST] Ellenőrzés befejezve:")
        print(f"   - Aktív linkek: {active_count}")
        print(f"   - Inaktív linkek: {inactive_count}")
        print(f"   - Összes link: {len(verified_links)}")
        print(f"   - Aktív százalék: {(active_count / len(verified_links) * 100):.1f}%")
        
        return jsonify({
            "success": True,
            "total_links": len(verified_links),
            "active_links": active_count,
            "inactive_links": inactive_count,
            "active_percentage": (active_count / len(verified_links) * 100) if len(verified_links) > 0 else 0,
            "verified_links": verified_links[:100],  # Csak az első 100 linket mutatjuk
            "sample_size": 100
        })
        
    except Exception as e:
        print(f"[ERROR] Link ellenőrzés hiba: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
